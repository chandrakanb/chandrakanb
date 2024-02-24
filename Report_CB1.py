import openpyxl

def apply_filter(sheet, keyword, column):
    # Apply filter for the keyword in the specified column
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    # Iterate through each cell in column B (from B2 to the last row)
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        # Check if the corresponding cell in column C has a value
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            # Set the value of the current cell in column B
            cell[0].value = value

# Open the Excel file
workbook = openpyxl.load_workbook('Execution_Report.xlsx')

# Select the active sheet
sheet = workbook.active

# Insert a column after the 1st column and name it "Script Type"
sheet.insert_cols(2)
sheet.cell(row=1, column=2).value = "Script Type"

# Apply filter for "KPIT" in the 3rd row
apply_filter(sheet, "KPIT", "C")

# Update column B for "KPIT"
update_column_B(sheet, "KPIT", "Test Script")

# Remove the filter
sheet.auto_filter.ref = None

# Apply filter for "pre" in the 3rd row
apply_filter(sheet, "pre", "C")

# Update column B for "pre"
update_column_B(sheet, "pre", "Precondition")

# Remove the filter
sheet.auto_filter.ref = None

# Apply filter for "post" in the 3rd row
apply_filter(sheet, "post", "C")

# Update column B for "post"
update_column_B(sheet, "post", "Postcondition")

# Remove the filter
sheet.auto_filter.ref = None

# Save the changes
workbook.save('Execution_Report.xlsx')
