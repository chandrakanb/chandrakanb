import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

def create_pivot_table(excel_file_path):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file_path)
        
        # Get the active sheet
        sheet = wb.active
        
        # Create a new sheet for the pivot table
        pivot_sheet = wb.create_sheet(title="temp")
        
        # Create a pivot table
        pivot, pivot_rows, pivot_columns = pivot_table(sheet)
        
        # Add pivot table to the new sheet
        for r_idx, row in enumerate(pivot_rows, 1):
            for c_idx, value in enumerate(row, 1):
                pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Create a new sheet for the transposed pivot table
        transposed_pivot_sheet = wb.create_sheet(title="Pivot_Table")
        
        # Transpose the pivot table
        transposed_pivot = pivot.transpose()
        
        # Add transposed pivot table to the new sheet
        for r_idx, row in enumerate(dataframe_to_rows(transposed_pivot, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                transposed_pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Remove unnecessary rows
        transposed_pivot_sheet.delete_rows(1, 2)
        
        # Remove the "temp" sheet
        wb.remove(wb["temp"])
        
        # Save the workbook with the specified name format
        output_file_name = f"Execution_Report_{pd.Timestamp.now().strftime('%dth_%B')}.xlsx"
        wb.save(output_file_name)
        
        print(f"Created the pivot table and transposed pivot table, saved as '{output_file_name}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def pivot_table(sheet):
    data = sheet.values
    cols = next(data)
    data = list(data)
    idx = [col for col in cols if col == "Script Type"]
    status_index = [col for col in cols if col == "Status"]
    script_type_index = cols.index("Script Type")

    # Get indices for columns
    idx_idx = cols.index(idx[0])
    status_idx = cols.index(status_index[0])
    script_type_idx = cols.index("Script Type")
    
    # Initialize dictionary to store counts
    counts = {}
    for row in data:
        if row[script_type_idx] not in counts:
            counts[row[script_type_idx]] = {}
        if row[status_idx] not in counts[row[script_type_idx]]:
            counts[row[script_type_idx]][row[status_idx]] = 0
        counts[row[script_type_idx]][row[status_idx]] += 1
    
    # Create dataframe from counts
    df = pd.DataFrame(counts).fillna(0)
    
    # Add Grand Total row and column
    df.loc['Grand Total'] = df.sum()
    df['Grand Total'] = df.sum(axis=1)
    
    # Reset index to make it a regular column
    df.reset_index(inplace=True)
    
    return df, dataframe_to_rows(df, index=False, header=True), cols

if __name__ == "__main__":
    excel_file_path = "Execution_Report.xlsx"
    create_pivot_table(excel_file_path)
