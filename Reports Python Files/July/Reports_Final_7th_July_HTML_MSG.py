import os
import re
import bs4
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import webbrowser
import pyautogui
import pyperclip
import requests
import base64

def version_details():
    feature = [
        "Creates Folder",
        "Generate Excel Report",
        "Adds \"Script Type\" Column to Excel Report",
        "Creates Separate Sheet for \"DryRun2.3\" and \"DryRun2.4\"",
        "Creates Pivot Table for \"DryRun2.3\" and \"DryRun2.4\"",
        "Takes Execution Summary Screenshot",
        "Renames Files",
        "Copies Files",
        "Copies Message to Clipboard"
    ]
    print("|\tVersion Details:\n|")
    for i, j in enumerate(feature, 1):
        print(f"|\t{i}.\t{j}.")

def read_html_file(execution_report_details):
    html_file = os.path.join(execution_report_details['execution_report_path'], execution_report_details['html_file'])
    df_summary = pd.read_html(html_file)
    df_summary = df_summary[0]
    df_summary = df_summary.filter(df_summary.columns[:2])
    infile = open(html_file)
    lines = infile.readlines()
    FileLines = "".join(lines)
    soup = BeautifulSoup(FileLines, 'html.parser')
    reportframe_add = soup.find('div', attrs={'id':'reportframe'})
    content_src = None
    for content in reportframe_add.contents:
        if isinstance(content, bs4.Tag):
            content_src= content['src']
            break
    content_src = content_src.replace("../","")
    content_src_paths = content_src.split("/")
    curr_path = execution_report_details['execution_report_path']
    curr_path = os.path.normpath(curr_path + os.sep + os.pardir)
    
    curr_path = os.path.join(curr_path,'XMLReport','Segments','XMLReportSegment_0.xml')
    
    tree = ET.parse(curr_path)
    root = tree.getroot()
    mainlist = list()
    for tcs_node in tree.iter("TestCases"):
        for tc_node in tcs_node.iter("TestCase"):
            tc_id = tc_node.find("testCaseID").text
            testCaseObjective = tc_node.find("testCaseObjective").text
            start_time = tc_node.find("startTime").text
            video_logger_link = tc_node.find("VideoLogLink").text
            result = tc_node.find("result").text
            total_time = tc_node.find("totalExecutionTime").text
            end_time = tc_node.find("endTime").text
            
            maindict = {"Test ID": tc_id,"Test Objective":testCaseObjective,"Start Time":start_time,"Video Logger Link":video_logger_link,"Status":result,"Total Time":total_time,"End Time":end_time}
            mainlist.append(maindict)
            
    df_detail_report = pd.DataFrame(mainlist)

    df_detail_report['file_path'] = curr_path
    return df_summary, df_detail_report

def iterate_tuple(dir_tuple):
    for item_tulpe in dir_tuple:
        if isinstance(item_tulpe, tuple):
            iterate_tuple(item_tulpe)
        elif isinstance(item_tulpe, list):
            for html_file in item_tulpe:
                if html_file == "HTMLReport.html":
                    execution_report_path = dir_tuple[0]
                    execution_report_name = execution_report_path.replace(os.getcwd(), "")

                    sub_folder_list = execution_report_name.split(os.sep)
                    test_pack = Iteration_count = None
                    for each_folder in sub_folder_list:
                        if "ITERATION" in each_folder.upper():
                            Iteration_count = each_folder
                            x = re.search("Iteration_[0-9]+", each_folder)
                            Iteration_count = x.group()
                            test_pack = each_folder.replace("_{}".format(Iteration_count),"")
                    Bench_name = sub_folder_list[1]
                    execution_report_name = sub_folder_list[2]
                    return {"execution_report_name": execution_report_name,
                            "execution_report_path": execution_report_path, "Bench_name":Bench_name,
                            "Iteration_count": Iteration_count,
                            "Test_Pack": test_pack,
                            "html_file": html_file}
    return

def add_additional_columns(df, execution_report_details):
    for col_ in ['Bench_name', 'execution_report_name', 'Test_Pack', 'Iteration_count']:
        df[col_] = execution_report_details[col_]
    return df

def generate_Detailed_Report_table(df):
    df_1 = df.filter(["Test_Pack", "Test ID"])
    df['Unique_col'] = df['Test_Pack'] + df['Test ID']
    df['Total Time'] = [datetime.strptime(i, "%H:%M:%S") for i in df['Total Time']]
    df_2 = pd.pivot_table(df, index=["Test_Pack", "Test ID"], columns='Status', aggfunc='count', values='Unique_col')
    return df_2

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            cell[0].value = value

def create_folder(input_string):
    os.makedirs(folder_name, exist_ok=True)
    print(f"|\tFolder created Successfully.")
    print('|' + '_' * 70 + '\n|\t')

def copy_pdf_files(input_string):
    pdf_reports_dir = os.path.join("XMLReports", "PdfReports")
    shutil.copy(os.path.join(pdf_reports_dir, "DashboardReport.pdf"), os.path.join(folder_name, f"{folder_name}_DashboardReport.pdf"))
    shutil.copy(os.path.join(pdf_reports_dir, "DetailedReport.pdf"), os.path.join(folder_name, f"{folder_name}_DetailedReport.pdf"))
    print(f"|\tPDF files copied Successfully.")
    print('|' + '_' * 70 + '\n|\t')
    
def take_screenshot(html_file_path, screenshot_path, crop_coordinates):
    # Construct the file URL using 'file://' protocol
    file_url = 'file:///' + os.path.abspath(html_file_path).replace('\\', '/')

    # Open the file URL in a web browser (opens in a new tab or window)
    webbrowser.open_new_tab(file_url)
    time.sleep(5)  # Adjust as necessary to ensure the page loads completely

    # Take a screenshot of the entire screen
    screenshot = pyautogui.screenshot()

    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(crop_coordinates)

    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(screenshot_path)
    print(f"|\tScreenshot Saved Successfully.")
    print('|' + '_' * 70 + '\n|\t')
    
    # Close the browser tab
    pyautogui.hotkey('ctrl', 'w')  # Close the tab (Ctrl + W)
    
    # Switch to the previous application
    pyautogui.hotkey('alt', 'tab')
    time.sleep(1)  # Wait for a second to ensure the tab is active

def get_date(input_date):
    # Convert number to string
    input_date_str = str(input_date)
    
    # Determine suffix based on last digit
    suffix = 'st' if input_date_str.endswith('1') and not input_date_str.endswith('11') else \
         'nd' if input_date_str.endswith('2') and not input_date_str.endswith('12') else \
         'rd' if input_date_str.endswith('3') and not input_date_str.endswith('13') else \
         'th'
         
    # Return ordinal string
    return input_date_str + suffix

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def extract_html_content(file_path):
    # Read the HTML file
    with open(file_path, "r", encoding="utf-8") as file:
        content = file.read()

    # Parse the HTML content
    soup = BeautifulSoup(content, "html.parser")

    # Extract the <head> part
    head = soup.head

    # Extract the specific <body> part
    summary_table_div = soup.find("div", id="div-summaryTable")
    pie_chart_div = soup.find("div", id="pieChartDiv")

    # Create a new BeautifulSoup object for the extracted content
    extracted_content = BeautifulSoup("<html></html>", "html.parser")
    extracted_html = extracted_content.html

    # Append the extracted head to the new BeautifulSoup object
    if head:
        extracted_html.append(head.extract())

    # Create and append the body with the extracted parts
    body = extracted_content.new_tag("body")
    extracted_html.append(body)

    # Create a container div to hold the summary table and pie chart side by side
    container_div = extracted_content.new_tag("div", **{"class": "container-fluid mt-5"})

    # Create a row div
    row_div = extracted_content.new_tag("div", **{"class": "row testPackData"})

    # Create the first column for the summary table
    summary_col_div = extracted_content.new_tag("div", **{"class": "col-sm-6"})
    if summary_table_div:
        summary_col_div.append(summary_table_div.extract())

    # Create the second column for the pie chart
    chart_col_div = extracted_content.new_tag("div", **{"class": "col-sm-6"})
    if pie_chart_div:
        chart_col_div.append(pie_chart_div.extract())

    # Append the columns to the row
    row_div.append(summary_col_div)
    row_div.append(chart_col_div)

    # Append the row to the container
    container_div.append(row_div)

    # Append the container to the body
    body.append(container_div)

    # Convert the extracted content to a string and return it
    extracted_html_str = str(extracted_content)
    
    print("|\tSummary table and Pie chart is saved Successfully.")
    print('|' + '_' * 70 + '\n|\t')
    
    return extracted_html_str

def send_message_to_teams(teams_webhook_url, message):
    try:
        # Prepare the request headers
        headers = {'Content-Type': 'application/json'}

        # Prepare the payload with the message
        payload = {
            "text": message,
            "attachments": [
                {
                    "contentType": "image/png",
                    "contentUrl": f"data:image/png;base64,{image_data}",
                    "name": f"{folder_name}_Execution_Summary.png"
                }
            ]
        }

        # Send the request
        response = requests.post(teams_webhook_url, headers=headers, json=payload)

        # Check for success
        if response.status_code == 200:
            print("Message sent successfully to Microsoft Teams!")
        else:
            print(f"Failed to send message. Status code: {response.status_code}, Response: {response.text}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    print('\n|' + '‾' * 70)
    #version_details()
    #print('|' + '_' * 70 + '\n|\t')
    
    # Get the currently active window title before opening the HTML file
    active_window_title_before = pyautogui.getActiveWindow().title
    
    # Parameters for creating folder and copying files
    result = "Result"
    
    #Teams webhook URL
    teams_webhook_url = 'https://kpitc.webhook.office.com/webhookb2/0e77225f-5e51-4ff3-bddd-87d2f5da5fc1@3539451e-b46e-4a26-a242-ff61502855c7/IncomingWebhook/bb555a09021747a696f5702d10233e92/fbaa0227-8d16-4af9-aaf0-17c8889162fe'
    
    # Take DryRun2.x from User
    flag = int(input("|\tDryRun2.x:\n|\t\t1.DryRun2.1\n|\t\t2.DryRun2.2\n|\t\t3.DryRun2.3\n|\t\t4.DryRun2.4\n|\tSelect DryRun2.x: "))
    cycle = 'DryRun2.1' if flag == 1 else 'DryRun2.2' if flag == 2 else 'DryRun2.3' if flag == 3 else 'DryRun2.4' if flag == 4 else None
    print('|' + '_' * 70 + '\n|\t')
    
    # Take Build Type from User
    input_build = int(input("|\tBuild Type:\n|\t\t1.Dev\n|\t\t2.Rel\n|\t\t3.Skip\n|\tSelect Build Type:"))
    build = "Dev_Build" if input_build == 1 else "Rel_Build" if input_build == 2 else None
    print('|' + '_' * 70 + '\n|\t')
    
    # Take Date from User and add correct suffix
    input_date = int(input("|\tEnter Date: "))
    date = get_date(input_date)
    print('|' + '_' * 70 + '\n|\t')

    # Take Month from User
    input_month = input("|\tEnter month: ")
    month = input_month.capitalize()
    print('|' + '_' * 70 + '\n|\t')
    
    # Take Execution Type from User
    overnight = "Overnight_Execution"
    flashing = "Auto_Flashing"
    flag = input("|\tExecution:\n|\t\t1. Overnight Execution\n|\t\t2. Auto Flashing\n|\tEnter your choice (1 or 2): ")
    execution = overnight if flag == "1" else flashing if flag == "2" else None
    print('|' + '_' * 70 + '\n|\t')
    
    # Take Bench Number from User
    input_bench = int(input("|\tEnter Bench Number: "))
    bench = (
        f"Bench0{input_bench}_T_Variant" if input_bench == 3 else
        f"Bench0{input_bench}_R_Variant" if input_bench == 6 else
        f"Bench0{input_bench}_TYAW" if input_bench == 9 else
        f"Bench{input_bench}_3BMA" if input_bench == 10 else
        f"Bench0{input_bench}_Q_Variant"
    )

    print('|' + '_' * 70 + '\n|\t')
    
    # Construct input_string
    parameters = [cycle, build, execution, date, month, bench] if input_build in [1, 2] else [execution, date, month, bench]
    input_string = '_'.join(parameters) + '_'
    
    # Folder name
    folder_name = input_string.strip("_").replace(" ", "_")
    
    # Create folder
    create_folder(input_string)
    
    # Take Execution_Summary.png
    html_file_path = 'MainDetailedReport.html'
    screenshot_path = os.path.join(folder_name, f"{folder_name}_Execution_Summary.png")
    crop_coordinates = (90, 160, 1350, 700) # crop coordinates: (left, upper, right, lower)
    
    # Call Screenshot Function 
    take_screenshot(html_file_path, screenshot_path, crop_coordinates)

    # Read Excel and HTML files
    df_summary_main = pd.DataFrame([])
    df_detail_report_main = pd.DataFrame([])

    for dir_tuple in os.walk(os.getcwd()):
        if iterate_tuple(dir_tuple):
            execution_report_details = iterate_tuple(dir_tuple)
            try:
                df_summary, df_detail_report = read_html_file(execution_report_details)
                df_summary = df_summary.T
                df_summary.columns = df_summary.loc['Summary'].to_list()
                df_summary = add_additional_columns(df_summary, execution_report_details)
                df_detail_report = add_additional_columns(df_detail_report, execution_report_details)
                df_summary_main = pd.concat([df_summary_main, df_summary], axis=0) if not df_summary_main.empty else df_summary
                df_detail_report_main = pd.concat([df_detail_report_main, df_detail_report], axis=0) if not df_detail_report_main.empty else df_detail_report
            except ValueError:
                print(execution_report_details)

    df_summary_main = df_summary_main.loc[df_summary_main['Start Time'] != 'Start Time']
    df_detail_report_main['Start Time_1'] = [datetime.strptime(i, "%d %b %Y %H:%M:%S") for i in df_detail_report_main['Start Time']]
    df_detail_report_main['Time_1'] = [datetime.strptime(i, "%H:%M:%S") for i in df_detail_report_main['Total Time']]
    End_Time = []
    for x, y in zip(df_detail_report_main['Start Time_1'], df_detail_report_main['Time_1']):
        __day__ = 0 if y.year == 1900 else y.day
        End_Time.append(x + timedelta(days=__day__, hours=y.hour, minutes=y.minute, seconds=y.second))
    df_detail_report_main['End Time'] = End_Time

    df_detail_report_main = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name",
                                                        "Iteration_count", "file_path"])

    df_detail_report_main_table = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name"])

    df_detail_report_main_table['Remark Failure Reason'] = ""

    # Write to Excel
    excel_file_path = os.path.join(folder_name, f"{folder_name}_Execution_Report.xlsx")
    writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')
    df_detail_report_main_table.to_excel(writer, sheet_name='Detailed Report', index=False)
    writer._save()
    
    #Define the border style and bold text
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    
    # Process Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    sheet.insert_cols(2)
    cell = sheet.cell(row=1, column=2)
    cell.value = "Script Type"
    cell.font = Font(bold=True)
    cell.border = thin_border
    apply_filter(sheet, "KPIT", "C")
    update_column_B(sheet, "KPIT", "Test Script")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "pre", "C")
    update_column_B(sheet, "pre", "Precondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "post", "C")
    update_column_B(sheet, "post", "Postcondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "auto", "C")
    update_column_B(sheet, "auto", "Precondition")
    sheet.auto_filter.ref = None
    workbook.save(excel_file_path)
    
    # Load the Excel workbook and the sheet with the result data
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    # Apply filter and create new sheets, including the first column (headers) in bold and bordered
    for version in ["2.3", "2.4"]:
        ws.auto_filter.ref = ws.dimensions
        apply_filter(ws, version, 'A')
        wb.create_sheet(f"DryRun{version}")
        new_sheet = wb[f"DryRun{version}"]
        
        # Copy headers
        for cell in ws[1]:
            new_sheet[cell.coordinate].value = cell.value
            new_sheet[cell.coordinate].font = bold_font
            new_sheet[cell.coordinate].border = thin_border

        # Copy filtered rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if version in str(row[0]):
                new_sheet.append(row)
        ws.auto_filter.ref = None
        
        # Create pivot table structure in new summary sheet
        summary_sheet_name = f"DryRun{version} Summary"
        wb.create_sheet(summary_sheet_name)
        summary_sheet = wb[summary_sheet_name]

        # Add headers to summary sheet
        headers = ["Script Type", "Fail", "Pass", "Verify Manually", "Total"]
        for col_num, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = bold_font
            #cell.border = thin_border

        # Prepare data for pivot table
        data = new_sheet.values
        df = pd.DataFrame(data)
        df.columns = df.iloc[0]
        df = df[1:]
        pivot = pd.pivot_table(df, index=["Script Type"], columns=["Status"], aggfunc='size', fill_value=0).reset_index()

        # Convert columns to numeric to avoid type errors
        for col in ["Fail", "Pass", "Verify Manually"]:
            if col in pivot.columns:
                pivot[col] = pd.to_numeric(pivot[col], errors='coerce').fillna(0).astype(int)
            else:
                pivot[col] = 0

        pivot['Total'] = pivot[["Fail", "Pass", "Verify Manually"]].sum(axis=1)
        
        # Reorder columns
        pivot = pivot[["Script Type", "Fail", "Pass", "Verify Manually", "Total"]]
        
        # Add total row
        total_row = pivot.sum(numeric_only=True)
        total_row["Script Type"] = "Total"
        pivot = pd.concat([pivot, total_row.to_frame().T], ignore_index=True)
        
        # Write pivot table to summary sheet
        for r_idx, row in enumerate(dataframe_to_rows(pivot, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = summary_sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.font = bold_font if r_idx == 1 else None
                cell.border = thin_border
        
        # Bold "Script Type" column
        for cell in summary_sheet["A:A"]:
            cell.font = bold_font
        
        # Generate HTML from pivot table
        if version == "2.3":
            HTML_2_3 = pivot.to_html(index=False)
        elif version == "2.4":
            HTML_2_4 = pivot.to_html(index=False)

        # Print pivot table
        print(f"Pivot Table for DryRun{version}:")
        print(pivot)
        print("\n")
    
    # Save the updated workbook
    wb.save(excel_file_path)
    wb.close()
    
    print(f"|\tExcel file Saved Successfully.")
    print('|' + '_' * 70 + '\n|\t')

    # Copy pdf files
    copy_pdf_files(input_string)
    
    # Call the function to extract Summary table and Pie chart from MainDetailedReport.html
    html_content = extract_html_content("MainDetailedReport.html")
    
    # Convert the image to base64
    with open(screenshot_path, "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    
    # Construct message_string
    if execution == overnight:
        date_month = f"{date} {month}, "
        execution_result = f"{execution.replace('_', ' ')} {result}:"
        flashing_result = f"{flashing.replace('_', ' ')} {result}:"
        DryRun = "DryRun2.3_"
        DryRun2 = "DryRun2.4_"
        if input_build in [1, 2]:
            message_string = f"{date_month}{cycle} {build.replace('_', ' ')} {execution.replace('_', ' ')} {bench.replace('_', ' ')}:\n{execution_result}\n{flashing_result}"
        else:
            message_string = f"{date_month}{execution.replace('_', ' ')} {bench.replace('_', ' ')}:\n{execution_result}\n{DryRun.replace('_', ' ')}{execution_result}\n{HTML_2_3}\n{DryRun2.replace('_', ' ')}{execution_result}\n{HTML_2_4}\n{flashing_result}"
        
        # Copy the output string to the clipboard
        pyperclip.copy(message_string)
        output_message = pyperclip.paste()
        if output_message == message_string:
            print("|\tMessage copied successfully.")
            print('|' + '_' * 70 + '\n|\t')
        
        # Call the function to send message to Teams
        send_message_to_teams(teams_webhook_url, message_string)

    # Switch back to the previously active window
    pyautogui.getWindowsWithTitle(active_window_title_before)[0].activate()
    
print("|\tAll tasks completed successfully.")
print('|' + '_' * 70 + '\n|\t')
input("|\tPress Enter to exit...")
print('|' + '_' * 70 + '\n')