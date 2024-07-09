import os
import re
import bs4
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import webbrowser
import pyautogui
import pyperclip

def read_html_file(execution_report_details):
    html_file = os.path.join(execution_report_details['execution_report_path'], execution_report_details['html_file'])
    df_summary = pd.read_html(html_file)
    df_summary = df_summary[0]
    df_summary = df_summary.filter(df_summary.columns[:2])
    infile = open(html_file)
    lines = infile.readlines()
    FileLines = "".join(lines)
    soup = BeautifulSoup(FileLines, 'html.parser')
    reportframe_add = soup.find('div', attrs={'id':'reportframe'})
    content_src = None
    for content in reportframe_add.contents:
        if isinstance(content, bs4.Tag):
            content_src= content['src']
            break
    content_src = content_src.replace("../","")
    content_src_paths = content_src.split("/")
    curr_path = execution_report_details['execution_report_path']
    curr_path = os.path.normpath(curr_path + os.sep + os.pardir)
    
    curr_path = os.path.join(curr_path,'XMLReport','Segments','XMLReportSegment_0.xml')
    
    tree = ET.parse(curr_path)
    root = tree.getroot()
    mainlist = list()
    for tcs_node in tree.iter("TestCases"):
        for tc_node in tcs_node.iter("TestCase"):
            tc_id = tc_node.find("testCaseID").text
            testCaseObjective = tc_node.find("testCaseObjective").text
            start_time = tc_node.find("startTime").text
            video_logger_link = tc_node.find("VideoLogLink").text
            result = tc_node.find("result").text
            total_time = tc_node.find("totalExecutionTime").text
            end_time = tc_node.find("endTime").text
            
            maindict = {"Test ID": tc_id,"Test Objective":testCaseObjective,"Start Time":start_time,"Video Logger Link":video_logger_link,"Status":result,"Total Time":total_time,"End Time":end_time}
            mainlist.append(maindict)
            
    df_detail_report = pd.DataFrame(mainlist)

    df_detail_report['file_path'] = curr_path
    return df_summary, df_detail_report


def iterate_tuple(dir_tuple):
    for item_tulpe in dir_tuple:
        if isinstance(item_tulpe, tuple):
            iterate_tuple(item_tulpe)
        elif isinstance(item_tulpe, list):
            for html_file in item_tulpe:
                if html_file == "HTMLReport.html":
                    execution_report_path = dir_tuple[0]
                    execution_report_name = execution_report_path.replace(os.getcwd(), "")

                    sub_folder_list = execution_report_name.split('\\')
                    test_pack = Iteration_count = None
                    for each_folder in sub_folder_list:
                        if "ITERATION" in each_folder.upper():
                            Iteration_count = each_folder
                            x = re.search("Iteration_[0-9]+", each_folder)
                            Iteration_count = x.group()
                            test_pack = each_folder.replace("_{}".format(Iteration_count),"")
                    Bench_name = sub_folder_list[1]
                    execution_report_name = sub_folder_list[2]
                    return {"execution_report_name": execution_report_name,
                            "execution_report_path": execution_report_path, "Bench_name":Bench_name,
                            "Iteration_count": Iteration_count,
                            "Test_Pack": test_pack,
                            "html_file": html_file}
    return


def add_additional_columns(df):
    for col_ in ['Bench_name', 'execution_report_name', 'Test_Pack', 'Iteration_count']:
        df[col_] = execution_report_details[col_]
    return df


def generate_Detailed_Report_table(df):
    df_1 = df.filter(["Test_Pack", "Test ID"])
    df['Unique_col'] = df['Test_Pack'] + df['Test ID']
    df['Total Time'] = [datetime.strptime(i, "%H:%M:%S") for i in df['Total Time']]
    df_2 = pd.pivot_table(df, index=["Test_Pack", "Test ID"], columns='Status', aggfunc='count', values='Unique_col')
    return df_2

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            cell[0].value = value

def create_folder_and_copy_files(input_string):
    folder_name = input_string.strip("_").replace(" ", "_")
    os.makedirs(folder_name, exist_ok=True)
    shutil.copy("Execution_Report.xlsx", os.path.join(folder_name, f"{folder_name}_Execution_Report.xlsx"))
    pdf_reports_dir = os.path.join("XMLReports", "PdfReports")
    shutil.copy(os.path.join(pdf_reports_dir, "DashboardReport.pdf"), os.path.join(folder_name, f"{folder_name}_DashboardReport.pdf"))
    shutil.copy(os.path.join(pdf_reports_dir, "DetailedReport.pdf"), os.path.join(folder_name, f"{folder_name}_DetailedReport.pdf"))
    
def take_screenshot(html_file_path, screenshot_path, crop_coordinates):
    # Construct the file URL using 'file://' protocol
    file_url = 'file:///' + os.path.abspath(html_file_path).replace('\\', '/')

    # Open the file URL in a web browser (opens in a new tab or window)
    webbrowser.open_new_tab(file_url)
    time.sleep(5)  # Adjust as necessary to ensure the page loads completely

    # Take a screenshot of the entire screen
    screenshot = pyautogui.screenshot()

    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(crop_coordinates)

    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(screenshot_path)
    print(f"|\tScreenshot Saved Successfully.")
    print('|' + '_' * 45 + '\n|\t')
    
    # Close the browser tab
    pyautogui.hotkey('ctrl', 'w')  # Close the tab (Ctrl + W)
    
    # Switch to the previous application
    pyautogui.hotkey('alt', 'tab')
    time.sleep(1)  # Wait for a second to ensure the tab is active

# Read Excel and HTML files
df_summary_main = pd.DataFrame([])
df_detail_report_main = pd.DataFrame([])

for dir_tuple in os.walk(os.getcwd()):
    if iterate_tuple(dir_tuple):
        execution_report_details = iterate_tuple(dir_tuple)
        try:
            df_summary, df_detail_report = read_html_file(execution_report_details)
            df_summary = df_summary.T
            df_summary.columns = df_summary.loc['Summary'].to_list()
            df_summary = add_additional_columns(df_summary)
            df_detail_report = add_additional_columns(df_detail_report)
            df_summary_main = pd.concat([df_summary_main,df_summary],axis=0) if df_summary_main.size > 0 else df_summary
            df_detail_report_main = pd.concat([df_detail_report_main,df_detail_report],axis=0) if df_detail_report_main.size > 0 else df_detail_report
        except ValueError:
            print(execution_report_details)

df_summary_main = df_summary_main.loc[df_summary_main['Start Time'] != 'Start Time']
df_detail_report_main['Start Time_1'] = [datetime.strptime(i, "%d %b %Y %H:%M:%S") for i in df_detail_report_main['Start Time']]
df_detail_report_main['Time_1'] = [datetime.strptime(i, "%H:%M:%S") for i in df_detail_report_main['Total Time']]
End_Time = list()
for x,y in zip(df_detail_report_main['Start Time_1'], df_detail_report_main['Time_1']):
    __day__ = 0 if y.year == 1900 else y.day
    End_Time.append(x + timedelta(days=__day__, hours=y.hour, minutes=y.minute, seconds=y.second))
df_detail_report_main['End Time'] = End_Time

df_detail_report_main = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name",
                                                    "Iteration_count", "file_path"])

df_detail_report_main_table = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name",])

df_detail_report_main_table['Remark Failure Reason'] = ""

# Write to Excel
writer = pd.ExcelWriter('Execution_Report.xlsx', engine='xlsxwriter')
df_detail_report_main_table.to_excel(writer, sheet_name='Detailed Report', index = False)
writer.close()

# Process Excel file
excel_file_path = "Execution_Report.xlsx"
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active
sheet.insert_cols(2)
sheet.cell(row=1, column=2).value = "Script Type"
apply_filter(sheet, "KPIT", "C")
update_column_B(sheet, "KPIT", "Test Script")
sheet.auto_filter.ref = None
apply_filter(sheet, "pre", "C")
update_column_B(sheet, "pre", "Precondition")
sheet.auto_filter.ref = None
apply_filter(sheet, "post", "C")
update_column_B(sheet, "post", "Postcondition")
sheet.auto_filter.ref = None
apply_filter(sheet, "Auto", "C")
update_column_B(sheet, "Auto", "Precondition")
sheet.auto_filter.ref = None
workbook.save(excel_file_path)

# Create folder and copy files
input_string = "DryRun2.3_DEV_Overnight_Execution_19th_June_Bench09_TYAW_"
create_folder_and_copy_files(input_string)

# Take Execution_Summary.png
html_file_path = 'MainDetailedReport.html'
screenshot_path = os.path.join(folder_name, f"{folder_name}_Execution_Summary.png")
crop_coordinates = (90, 160, 1350, 450) # crop coordinates: (left, upper, right, lower)

# Call Screenshot Function 
take_screenshot(html_file_path, screenshot_path, crop_coordinates)

os.remove("Execution_Report.xlsx")

print("All tasks completed successfully.")
