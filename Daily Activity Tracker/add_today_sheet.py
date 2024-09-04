import os
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

def copy_sheet_with_date(sheet_name_to_copy, excel_file_path):
    # Load the existing workbook
    book = load_workbook(excel_file_path)
    
    # Check if the sheet to copy exists
    if sheet_name_to_copy not in book.sheetnames:
        print(f"Sheet '{sheet_name_to_copy}' not found in {excel_file_path}.")
        return
    
    # Load the sheet to copy
    sheet_to_copy = book[sheet_name_to_copy]

    # Get today's date in the format "day_month"
    today = datetime.now()
    formatted_date = get_date_suffix(today)
    
    # Define the new sheet name
    new_sheet_name = f"{formatted_date}"
    
    # Copy the sheet
    new_sheet = book.copy_worksheet(sheet_to_copy)
    new_sheet.title = new_sheet_name
    
    # Save the workbook with the new sheet
    book.save(excel_file_path)
    
    print(f"New sheet '{new_sheet_name}' is added to '{excel_file_path}'.")

def get_date_suffix(date):
    day = date.day
    month = date.strftime("%B")  # Full month name
    
    # Determine the ordinal suffix
    if 10 <= day % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    
    formatted_date = f"{day}{suffix}_{month}"
    return formatted_date

if __name__ == "__main__":
    # Define the Excel file name and path
    excel_file_name = "Daily_Activity_Tracker_September.xlsx"
    excel_file_path = os.path.abspath(excel_file_name)
    
    # Name of the sheet to copy
    sheet_name_to_copy = "day_month"
    
    # Call the function to copy the sheet with the new date-based name
    copy_sheet_with_date(sheet_name_to_copy, excel_file_name)
