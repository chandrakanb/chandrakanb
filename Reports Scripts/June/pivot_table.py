import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def create_pivot_table(excel_file_path):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Read the data into a DataFrame
    data = sheet.values
    columns = next(data)[0:]  # first row is the header
    df = pd.DataFrame(data, columns=columns)

    # Debugging: Print the DataFrame info and head
    print("DataFrame info:")
    print(df.info())
    print("DataFrame head:")
    print(df.head())

    # Validate columns
    required_columns = ['Script Type', 'Status']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' is missing from the data")

    # Ensure 'Script Type' is 1-dimensional
    df['Script Type'] = df['Script Type'].apply(lambda x: x if isinstance(x, str) else str(x))

    # Debugging: Print unique values of 'Script Type' column
    print("Unique values in 'Script Type' column:")
    print(df['Script Type'].unique())

    # Create a pivot table
    pivot_table = pd.pivot_table(df, 
                                 index='Script Type', 
                                 columns='Status', 
                                 values='Script Type', 
                                 aggfunc='count', 
                                 fill_value=0)

    # Convert the pivot table to a DataFrame
    pivot_df = pd.DataFrame(pivot_table.to_records())

    # Create a new worksheet and write the pivot table to it
    new_sheet = workbook.create_sheet(title="Pivot Table")
    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            new_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    workbook.save(excel_file_path)

if __name__ == "__main__":
    # Define the file path
    cycle = "DryRun2.3"
    overnight = "Overnight"
    execution = "Execution"
    date = "1st"
    month = "June"
    bench = "Bench9"
    folder_name = f"{cycle}_{overnight}_{execution}_{date}_{month}_{bench}"
    excel_file_name = f"{folder_name}_Execution_Report.xlsx"
    excel_file_path = os.path.join(folder_name, excel_file_name)
    
    # Check if the file exists
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file '{excel_file_path}' not found.")
    
    # Create the pivot table
    create_pivot_table(excel_file_path)

print("Pivot table created and saved successfully.")
