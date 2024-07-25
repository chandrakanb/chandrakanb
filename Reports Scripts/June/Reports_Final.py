import os
import re
import bs4
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import webbrowser
import pyautogui
import pyperclip

def read_html_file(execution_report_details):
    html_file = os.path.join(execution_report_details['execution_report_path'], execution_report_details['html_file'])
    df_summary = pd.read_html(html_file)
    df_summary = df_summary[0]
    df_summary = df_summary.filter(df_summary.columns[:2])
    infile = open(html_file)
    lines = infile.readlines()
    FileLines = "".join(lines)
    soup = BeautifulSoup(FileLines, 'html.parser')
    reportframe_add = soup.find('div', attrs={'id':'reportframe'})
    content_src = None
    for content in reportframe_add.contents:
        if isinstance(content, bs4.Tag):
            content_src= content['src']
            break
    content_src = content_src.replace("../","")
    content_src_paths = content_src.split("/")
    curr_path = execution_report_details['execution_report_path']
    curr_path = os.path.normpath(curr_path + os.sep + os.pardir)
    
    curr_path = os.path.join(curr_path,'XMLReport','Segments','XMLReportSegment_0.xml')
    
    tree = ET.parse(curr_path)
    root = tree.getroot()
    mainlist = list()
    for tcs_node in tree.iter("TestCases"):
        for tc_node in tcs_node.iter("TestCase"):
            tc_id = tc_node.find("testCaseID").text
            testCaseObjective = tc_node.find("testCaseObjective").text
            start_time = tc_node.find("startTime").text
            video_logger_link = tc_node.find("VideoLogLink").text
            result = tc_node.find("result").text
            total_time = tc_node.find("totalExecutionTime").text
            end_time = tc_node.find("endTime").text
            
            maindict = {"Test ID": tc_id,"Test Objective":testCaseObjective,"Start Time":start_time,"Video Logger Link":video_logger_link,"Status":result,"Total Time":total_time,"End Time":end_time}
            mainlist.append(maindict)
            
    df_detail_report = pd.DataFrame(mainlist)

    df_detail_report['file_path'] = curr_path
    return df_summary, df_detail_report

def iterate_tuple(dir_tuple):
    for item_tulpe in dir_tuple:
        if isinstance(item_tulpe, tuple):
            iterate_tuple(item_tulpe)
        elif isinstance(item_tulpe, list):
            for html_file in item_tulpe:
                if html_file == "HTMLReport.html":
                    execution_report_path = dir_tuple[0]
                    execution_report_name = execution_report_path.replace(os.getcwd(), "")

                    sub_folder_list = execution_report_name.split(os.sep)
                    test_pack = Iteration_count = None
                    for each_folder in sub_folder_list:
                        if "ITERATION" in each_folder.upper():
                            Iteration_count = each_folder
                            x = re.search("Iteration_[0-9]+", each_folder)
                            Iteration_count = x.group()
                            test_pack = each_folder.replace("_{}".format(Iteration_count),"")
                    Bench_name = sub_folder_list[1]
                    execution_report_name = sub_folder_list[2]
                    return {"execution_report_name": execution_report_name,
                            "execution_report_path": execution_report_path, "Bench_name":Bench_name,
                            "Iteration_count": Iteration_count,
                            "Test_Pack": test_pack,
                            "html_file": html_file}
    return

def add_additional_columns(df, execution_report_details):
    for col_ in ['Bench_name', 'execution_report_name', 'Test_Pack', 'Iteration_count']:
        df[col_] = execution_report_details[col_]
    return df

def generate_Detailed_Report_table(df):
    df_1 = df.filter(["Test_Pack", "Test ID"])
    df['Unique_col'] = df['Test_Pack'] + df['Test ID']
    df['Total Time'] = [datetime.strptime(i, "%H:%M:%S") for i in df['Total Time']]
    df_2 = pd.pivot_table(df, index=["Test_Pack", "Test ID"], columns='Status', aggfunc='count', values='Unique_col')
    return df_2

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            cell[0].value = value

def create_folder(input_string):
    os.makedirs(folder_name, exist_ok=True)
    print(f"|\tFolder created Successfully.")
    print('|' + '_' * 45 + '\n|\t')

def copy_pdf_files(input_string):
    pdf_reports_dir = os.path.join("XMLReports", "PdfReports")
    shutil.copy(os.path.join(pdf_reports_dir, "DashboardReport.pdf"), os.path.join(folder_name, f"{folder_name}_DashboardReport.pdf"))
    shutil.copy(os.path.join(pdf_reports_dir, "DetailedReport.pdf"), os.path.join(folder_name, f"{folder_name}_DetailedReport.pdf"))
    print(f"|\tPDF files copied Successfully.")
    print('|' + '_' * 45 + '\n|\t')
    
def take_screenshot(html_file_path, screenshot_path, crop_coordinates):
    # Construct the file URL using 'file://' protocol
    file_url = 'file:///' + os.path.abspath(html_file_path).replace('\\', '/')

    # Open the file URL in a web browser (opens in a new tab or window)
    webbrowser.open_new_tab(file_url)
    time.sleep(5)  # Adjust as necessary to ensure the page loads completely

    # Take a screenshot of the entire screen
    screenshot = pyautogui.screenshot()

    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(crop_coordinates)

    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(screenshot_path)
    print(f"|\tScreenshot Saved Successfully.")
    print('|' + '_' * 45 + '\n|\t')
    
    # Close the browser tab
    pyautogui.hotkey('ctrl', 'w')  # Close the tab (Ctrl + W)
    
    # Switch to the previous application
    pyautogui.hotkey('alt', 'tab')
    time.sleep(1)  # Wait for a second to ensure the tab is active

def get_date(input_date):
    # Convert number to string
    input_date_str = str(input_date)
    
    # Determine suffix based on last digit
    if input_date_str.endswith('1') and not input_date_str.endswith('11'):
        suffix = 'st'
    elif input_date_str.endswith('2') and not input_date_str.endswith('12'):
        suffix = 'nd'
    elif input_date_str.endswith('3') and not input_date_str.endswith('13'):
        suffix = 'rd'
    else:
        suffix = 'th'
    
    # Return ordinal string
    return input_date_str + suffix

if __name__ == "__main__":
    print('\n|' + '‾' * 45)
    
    # Get the currently active window title before opening the HTML file
    active_window_title_before = pyautogui.getActiveWindow().title
    
    # Parameters for creating folder and copying files
    cycle = "DryRun2.3"
    build = "Build"
    result = "Result"

    # Take Build Type from User
    input_build = int(input("|\tBuild Type:\n|\t\t1.Dev\n|\t\t2.Rel\n|\tSelect Build Type:"))
    if input_build == 1:
        dev_or_rel = "Dev"
    elif input_build == 2:
        dev_or_rel = "Rel"
    print('|' + '_' * 45 + '\n|\t')
    
    # Take Date from User and add correct suffix
    input_date = int(input("|\tEnter Date: "))
    date = get_date(input_date)
    print('|' + '_' * 45 + '\n|\t')

    # Take Month from User
    input_month = input("|\tEnter month: ")
    month = input_month.capitalize()
    print('|' + '_' * 45 + '\n|\t')
    
    # Take Execution Type from User
    overnight = "Overnight_Execution"
    flashing = "Auto_Flashing"
    flag = input("|\tExecution:\n|\t\t1. Overnight Execution\n|\t\t2. Auto Flashing\n|\tEnter your choice (1 or 2): ")
    if flag == "1":
        execution = overnight
    elif flag == "2":
        execution = flashing
    print('|' + '_' * 45 + '\n|\t')
    
    # Take Bench Number from User
    input_bench = int(input("|\tEnter Bench Number: "))
    if input_bench == 3:
        bench = f"Bench0{input_bench}_T_Variant"
    elif input_bench == 6:
        bench = f"Bench0{input_bench}_R_Variant"
    elif input_bench == 9:
        bench = f"Bench0{input_bench}_TYAW"
    elif input_bench == 10:
        bench = f"Bench{input_bench}_3BMA"
    else:
        bench = f"Bench0{input_bench}_Q_Variant"
    print('|' + '_' * 45 + '\n|\t')
    
    # Construct input_string
    parameters = [cycle, dev_or_rel, build, execution, date, month, bench]
    input_string = '_'.join(parameters) + '_'
    
    # Folder name
    folder_name = input_string.strip("_").replace(" ", "_")
    
    # Create folder
    create_folder(input_string)
    
    # Take Execution_Summary.png
    html_file_path = 'MainDetailedReport.html'
    screenshot_path = os.path.join(folder_name, f"{folder_name}_Execution_Summary.png")
    crop_coordinates = (90, 160, 1350, 450) # crop coordinates: (left, upper, right, lower)
    
    # Call Screenshot Function 
    take_screenshot(html_file_path, screenshot_path, crop_coordinates)

    # Read Excel and HTML files
    df_summary_main = pd.DataFrame([])
    df_detail_report_main = pd.DataFrame([])

    for dir_tuple in os.walk(os.getcwd()):
        if iterate_tuple(dir_tuple):
            execution_report_details = iterate_tuple(dir_tuple)
            try:
                df_summary, df_detail_report = read_html_file(execution_report_details)
                df_summary = df_summary.T
                df_summary.columns = df_summary.loc['Summary'].to_list()
                df_summary = add_additional_columns(df_summary, execution_report_details)
                df_detail_report = add_additional_columns(df_detail_report, execution_report_details)
                df_summary_main = pd.concat([df_summary_main, df_summary], axis=0) if not df_summary_main.empty else df_summary
                df_detail_report_main = pd.concat([df_detail_report_main, df_detail_report], axis=0) if not df_detail_report_main.empty else df_detail_report
            except ValueError:
                print(execution_report_details)

    df_summary_main = df_summary_main.loc[df_summary_main['Start Time'] != 'Start Time']
    df_detail_report_main['Start Time_1'] = [datetime.strptime(i, "%d %b %Y %H:%M:%S") for i in df_detail_report_main['Start Time']]
    df_detail_report_main['Time_1'] = [datetime.strptime(i, "%H:%M:%S") for i in df_detail_report_main['Total Time']]
    End_Time = []
    for x, y in zip(df_detail_report_main['Start Time_1'], df_detail_report_main['Time_1']):
        __day__ = 0 if y.year == 1900 else y.day
        End_Time.append(x + timedelta(days=__day__, hours=y.hour, minutes=y.minute, seconds=y.second))
    df_detail_report_main['End Time'] = End_Time

    df_detail_report_main = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name",
                                                        "Iteration_count", "file_path"])

    df_detail_report_main_table = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name"])

    df_detail_report_main_table['Remark Failure Reason'] = ""

    # Write to Excel
    excel_file_path = os.path.join(folder_name, f"{folder_name}_Execution_Report.xlsx")
    writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')
    df_detail_report_main_table.to_excel(writer, sheet_name='Detailed Report', index=False)
    writer._save()

    # Process Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    sheet.insert_cols(2)
    sheet.cell(row=1, column=2).value = "Script Type"
    apply_filter(sheet, "KPIT", "C")
    update_column_B(sheet, "KPIT", "Test Script")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "pre", "C")
    update_column_B(sheet, "pre", "Precondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "post", "C")
    update_column_B(sheet, "post", "Postcondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "auto", "C")
    update_column_B(sheet, "auto", "Precondition")
    sheet.auto_filter.ref = None
    workbook.save(excel_file_path)
    print(f"|\tExcel file Saved Successfully.")
    print('|' + '_' * 45 + '\n|\t')

    # Copy pdf files
    copy_pdf_files(input_string)
    
    # Construct message_string
    if execution == overnight:
        date_month = [date, month]
        cycle_dev_or_rel_build_execution_bench = [cycle, dev_or_rel, build, execution, bench]
        execution_result = [execution, result]
        flashing_result = [flashing, result]
        output_string_1 = ' '.join(date_month)+', '+' '.join(cycle_dev_or_rel_build_execution_bench).replace("_", " ")+':'
        output_string_2 = ' '.join(execution_result).replace("_", " ")+':'
        output_string_3 = ' '.join(flashing_result).replace("_", " ")+':'
        message_string = f"{output_string_1}\n{output_string_2}\n{output_string_3}"
    
    
        # Copy the output string to the clipboard
        pyperclip.copy(message_string)
        output_message = pyperclip.paste()
        if output_message == message_string:
            print("|\tMessage copied successfully.")
            print('|' + '_' * 45 + '\n|\t')

    # Switch back to the previously active window
    pyautogui.getWindowsWithTitle(active_window_title_before)[0].activate()
    
print("|\tAll tasks completed successfully.")
print('|' + '_' * 45 + '\n')
input("Press Enter to exit...")