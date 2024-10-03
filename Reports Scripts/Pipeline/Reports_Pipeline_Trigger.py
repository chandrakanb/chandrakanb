import os
import re
import bs4
import sys
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import webbrowser
import pyautogui
import pyperclip
import requests
import base64
from io import StringIO
from html2image import Html2Image
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from PIL import Image
import subprocess
from contextlib import contextmanager, redirect_stdout, redirect_stderr
from pathlib import Path
from msedge.selenium_tools import Edge, EdgeOptions

def read_html_file(execution_report_details):
    html_file = os.path.join(execution_report_details['execution_report_path'], execution_report_details['html_file'])
    df_summary = pd.read_html(html_file)
    df_summary = df_summary[0]
    df_summary = df_summary.filter(df_summary.columns[:2])
    infile = open(html_file)
    lines = infile.readlines()
    FileLines = "".join(lines)
    soup = BeautifulSoup(FileLines, 'html.parser')
    reportframe_add = soup.find('div', attrs={'id':'reportframe'})
    content_src = None
    for content in reportframe_add.contents:
        if isinstance(content, bs4.Tag):
            content_src= content['src']
            break
    content_src = content_src.replace("../","")
    content_src_paths = content_src.split("/")
    curr_path = execution_report_details['execution_report_path']
    curr_path = os.path.normpath(curr_path + os.sep + os.pardir)
    
    curr_path = os.path.join(curr_path,'XMLReport','Segments','XMLReportSegment_0.xml')
    
    tree = ET.parse(curr_path)
    root = tree.getroot()
    mainlist = list()
    for tcs_node in tree.iter("TestCases"):
        for tc_node in tcs_node.iter("TestCase"):
            tc_id = tc_node.find("testCaseID").text
            testCaseObjective = tc_node.find("testCaseObjective").text
            start_time = tc_node.find("startTime").text
            video_logger_link = tc_node.find("VideoLogLink").text
            result = tc_node.find("result").text
            total_time = tc_node.find("totalExecutionTime").text
            end_time = tc_node.find("endTime").text
            
            maindict = {"Test ID": tc_id,"Test Objective":testCaseObjective,"Start Time":start_time,"Video Logger Link":video_logger_link,"Status":result,"Total Time":total_time,"End Time":end_time}
            mainlist.append(maindict)
            
    df_detail_report = pd.DataFrame(mainlist)

    df_detail_report['file_path'] = curr_path
    return df_summary, df_detail_report

def iterate_tuple(dir_tuple):
    for item_tulpe in dir_tuple:
        if isinstance(item_tulpe, tuple):
            iterate_tuple(item_tulpe)
        elif isinstance(item_tulpe, list):
            for html_file in item_tulpe:
                if html_file == "HTMLReport.html":
                    execution_report_path = dir_tuple[0]
                    execution_report_name = execution_report_path.replace(os.getcwd(), "")

                    sub_folder_list = execution_report_name.split(os.sep)
                    test_pack = Iteration_count = None
                    for each_folder in sub_folder_list:
                        if "ITERATION" in each_folder.upper():
                            Iteration_count = each_folder
                            x = re.search("Iteration_[0-9]+", each_folder)
                            Iteration_count = x.group()
                            test_pack = each_folder.replace("_{}".format(Iteration_count),"")
                    Bench_name = sub_folder_list[1]
                    execution_report_name = sub_folder_list[2]
                    return {"execution_report_name": execution_report_name,
                            "execution_report_path": execution_report_path, "Bench_name":Bench_name,
                            "Iteration_count": Iteration_count,
                            "Test_Pack": test_pack,
                            "html_file": html_file}
    return

def add_additional_columns(df, execution_report_details):
    for col_ in ['Bench_name', 'execution_report_name', 'Test_Pack', 'Iteration_count']:
        df[col_] = execution_report_details[col_]
    return df

def generate_Detailed_Report_table(df):
    df_1 = df.filter(["Test_Pack", "Test ID"])
    df['Unique_col'] = df['Test_Pack'] + df['Test ID']
    df['Total Time'] = [datetime.strptime(i, "%H:%M:%S") for i in df['Total Time']]
    df_2 = pd.pivot_table(df, index=["Test_Pack", "Test ID"], columns='Status', aggfunc='count', values='Unique_col')
    return df_2

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            cell[0].value = value

def create_folder(folder_path):
    """Create a directory if it does not exist."""
    os.makedirs(folder_path, exist_ok=True)
    print("Folder created successfully.")

def copy_pdf_files(Execution_Path, reports_prefix, folder_path):
    pdf_reports_dir = os.path.join(f"{Execution_Path}","XMLReports", "PdfReports")
    shutil.copy(os.path.join(pdf_reports_dir, "DashboardReport.pdf"), os.path.join(folder_path, f"{reports_prefix}_DashboardReport.pdf"))
    shutil.copy(os.path.join(pdf_reports_dir, "DetailedReport.pdf"), os.path.join(folder_path, f"{reports_prefix}_DetailedReport.pdf"))
    print(f"{"_".join(reports_prefix.split("_")[-2:]).replace("_", " ")} PDF files copied successfully.")

def take_screenshot(html_file_path, screenshot_path, crop_coordinates, reports_prefix):
    """Take a screenshot of the specified HTML file and save it."""
    # Set up Edge options for headless mode
    options = Options()
    options.use_chromium = True
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    
    # Construct the file URL using 'file://' protocol
    file_url = 'file:///' + os.path.abspath(html_file_path).replace('\\', '/')
    
    # Set up the Edge WebDriver with headless options
    driver = webdriver.Edge(options=options)
    
    # Open the HTML file in the headless browser
    driver.get(file_url)
    
    # Take a screenshot of the entire page
    screenshot_data = driver.get_screenshot_as_png()
    
    # Save the screenshot to a temporary file
    temp_screenshot_path = 'temp_screenshot.png'
    with open(temp_screenshot_path, 'wb') as f:
        f.write(screenshot_data)
    
    # Open the screenshot image
    screenshot = Image.open(temp_screenshot_path)
    
    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(crop_coordinates)
    
    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(screenshot_path)
    print(f"{"_".join(reports_prefix.split("_")[-2:]).replace("_", " ")} Screenshot saved successfully.")
    
    # Clean up and close the browser
    driver.quit()
    os.remove(temp_screenshot_path)

def get_execution_date_month(html_file_path):
    """
    This function reads an HTML report file to extract the start time of the execution.
    It then parses the start time to obtain the date, month, year, hours, and minutes.
    The date is adjusted if the hour is between 12 midnight and 12 noon of the next day.
    Finally, it converts the abbreviated month name to the full month name and returns the date and month.
    """
    # Read the content of the HTML file
    with open(html_file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the start time
    start_time_tag = soup.find('th', string='Start Time')
    start_time_value = start_time_tag.find_next_sibling('td').text if start_time_tag else None

    # Split the start time into date, month, year, hours, and minutes
    date_parts = start_time_value.split(' ')
    date_digit = int(date_parts[0])  # "date" as an integer
    month_abbr = date_parts[1]  # "month"
    year = date_parts[2]  # "year"
    time_parts = date_parts[3].split(':')
    time_hh = time_parts[0]  # "hours"
    time_mm = time_parts[1]  # "minutes"

    # Adjust the date based on the hour
    # If hours is between 12 midnight and next day's 12 noon, then date minus one
    if int(time_hh) <= 12:
        date_digit -= 1
    
    # Add correct suffix to the date based on last digit
    date = get_date(date_digit)
    
    # Convert abbreviated month to full month name
    month = get_month(month_abbr)

    return date, month, year

def get_date(date_digit):
    """Return the date with the appropriate ordinal suffix."""
    date_str = str(date_digit)
    suffix = 'st' if date_str.endswith('1') and not date_str.endswith('11') else \
             'nd' if date_str.endswith('2') and not date_str.endswith('12') else \
             'rd' if date_str.endswith('3') and not date_str.endswith('13') else \
             'th'
    return date_str + suffix

def get_month(month_abbr):
    """Return the full month name based on abbreviated month names."""
    months = {
        "Jan": "January", "Feb": "February", "Mar": "March",
        "Apr": "April", "May": "May", "Jun": "June",
        "Jul": "July", "Aug": "August", "Sep": "September",
        "Oct": "October", "Nov": "November", "Dec": "December"
    }
    return months.get(month_abbr, month_abbr)

def get_bench(input_bench):
    """Return the bench name with the appropriate variant name."""
    bench = (
        f"Bench0{input_bench}_T_Variant" if input_bench == 7 else
        f"Bench0{input_bench}_R_Variant" if input_bench == 6 else
        f"Bench0{input_bench}_TYAW" if input_bench == 9 else
        f"Bench{input_bench}_3BMA" if input_bench == 10 else
        f"Bench0{input_bench}_Q_Variant"
    )
    bench_abr = bench.split('_')[0]
    return bench, bench_abr

def create_zip_folder(folder_name):
    zip_filename = f"{folder_name}.zip"
    shutil.make_archive(folder_name, 'zip', folder_name)
    print("ZIP folder created successfully.")
    return zip_filename

def send_message_with_zip_to_teams(teams_webhook_url, message_string, message_string_1, message_string_2, overnight_execution_screenshot_path, auto_flashing_screenshot_path):
    try:
        # Prepare the request headers
        headers = {'Content-Type': 'application/json'}
        
        # Prepare the image attachments
        with open(overnight_execution_screenshot_path, "rb") as f:
            encoded_overnight_execution_screenshot = base64.b64encode(f.read()).decode('utf-8')
        
        with open(auto_flashing_screenshot_path, "rb") as f:
            encoded_auto_flashing_screenshot = base64.b64encode(f.read()).decode('utf-8')
        
        # Construct the adaptive card content
        adaptive_card_content = [
            {
                "type": "TextBlock",
                "text": message_string,
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": message_string_1,
                "wrap": True
            },
            {
                "type": "Image",
                "url": f"data:image/png;base64,{encoded_overnight_execution_screenshot}",
                "size": "Auto"  # Set to "Auto" to maintain original size and high quality
            },
            {
                "type": "TextBlock",
                "text": message_string_2,
                "wrap": True
            },
            {
                "type": "Image",
                "url": f"data:image/png;base64,{encoded_auto_flashing_screenshot}",
                "size": "Auto"  # Set to "Auto" to maintain original size and high quality
            }
        ]

        # Prepare the adaptive card payload with the image and tables
        payload = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "content": {
                        "type": "AdaptiveCard",
                        "version": "1.2",
                        "body": adaptive_card_content,
                        "msteams": {
                            "width": "Full"
                        }
                    }
                }
            ]
        }

        # Send the request
        response = requests.post(teams_webhook_url, headers=headers, json=payload)

        # Check for success
        if response.status_code == 200:
            print("Message sent successfully to Microsoft Teams!")
        else:
            print(f"Failed to send message. Status code: {response.status_code}, Response: {response.text}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    # Get the currently active window title 
    #active_window_title_before = pyautogui.getActiveWindow().title
    
    #Teams webhook URL
    teams_webhook_url = 'https://kpitc.webhook.office.com/webhookb2/0e77225f-5e51-4ff3-bddd-87d2f5da5fc1@3539451e-b46e-4a26-a242-ff61502855c7/IncomingWebhook/bb555a09021747a696f5702d10233e92/fbaa0227-8d16-4af9-aaf0-17c8889162fe'

    # Check if enough arguments are provided
    if len(sys.argv) > 2:
        # Take paths for overnight execution and auto flashing from user arguments
        Overnight_Execution_Path = Path(sys.argv[1])
        Auto_Flashing_Path = Path(sys.argv[2])
    else:
        print("Error: Not enough arguments provided.")
        print("Usage: python script_name.py <Overnight_Execution_Path> <Auto_Flashing_Path>")

    # Define html file locations 
    overnight_execution_html_file_path = os.path.join(Overnight_Execution_Path, "MainDetailedReport.html")
    auto_flashing_html_file_path = os.path.join(Auto_Flashing_Path, "MainDetailedReport.html")

    # Set the execution type
    execution = "Overnight_Execution"
    auto_flashing = "Auto_Flashing"
    # execution_abr = 

    # Provide the full path to msedgedriver.exe
    # driver_path = os.path.abspath(f"msedgedriver.exe")

    # Get the execution date and month from the HTML report
    date, month, year = get_execution_date_month(overnight_execution_html_file_path)

    # Prompt the user to enter the bench number
    input_bench = int(sys.argv[3])    # input("Enter Bench Number: "))
    bench, bench_abr = get_bench(input_bench)

    # Construct folder_name
    parameters = [bench_abr, date, month, execution]
    folder_name = '_'.join(parameters).strip("_").replace(" ", "_")
    execution_reports_prefix = '_'.join(parameters).strip("_").replace(" ", "_")
    parameters_1 = [bench_abr, date, month, auto_flashing]
    auto_flashing_reports_prefix = '_'.join(parameters_1).strip("_").replace(" ", "_")

    # Create folder
    reports_folder = "reports"
    os.makedirs(reports_folder, exist_ok=True)
    folder_path = os.path.join("reports", f"{folder_name}")
    create_folder(folder_path)
    
    # Copy pdf files for Overnight Execution
    copy_pdf_files(Overnight_Execution_Path, execution_reports_prefix, folder_path)
    
    # Copy pdf files for Auto Flashing
    copy_pdf_files(Auto_Flashing_Path, auto_flashing_reports_prefix, folder_path)
    
    # Take Execution_Summary.png
    overnight_execution_screenshot_path = os.path.join(folder_path, f"{execution_reports_prefix}_Summary_Table_Pie_Chart.png")
    auto_flashing_screenshot_path = os.path.join(folder_path, f"{auto_flashing_reports_prefix}_Summary_Table_Pie_Chart.png")
    # Bench Crop Coordinates
    crop_coordinates = (90, 63, 1365, 341) # crop coordinates: (left, upper, right, lower)
    # Lappy Crop Coordinates
    # crop_coordinates = (115, 80, 1695, 425) # crop coordinates: (left, upper, right, lower)

    # Call Screenshot Function for Auto-Flashing
    take_screenshot(overnight_execution_html_file_path, overnight_execution_screenshot_path, crop_coordinates, execution_reports_prefix)

    # Call Screenshot Function for Overnight Execution
    take_screenshot(auto_flashing_html_file_path, auto_flashing_screenshot_path, crop_coordinates, auto_flashing_reports_prefix)

    #Define the border style and bold text
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    # Process Excel file
    # Write to Excel
    excel_file_path = os.path.join(Overnight_Execution_Path, "DRT_Execution_Report.xlsx")
    # folder_excel_file_path = os.path.join(folder_path, f"DRT_Execution_Report.xlsx")
    # shutil.copy(excel_file_path, folder_excel_file_path)
    pivot_excel_file_path = os.path.join(folder_path, f"{execution_reports_prefix}_Execution_Report.xlsx")
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    sheet.insert_cols(2)
    cell = sheet.cell(row=1, column=2)
    cell.value = "Script Type"
    cell.font = Font(bold=True)
    cell.border = thin_border
    apply_filter(sheet, "KPIT", "C")
    update_column_B(sheet, "KPIT", "Test Script")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "post", "C")
    update_column_B(sheet, "post", "Postcondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "pre", "C")
    update_column_B(sheet, "pre", "Precondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "autof", "C")
    update_column_B(sheet, "autof", "Precondition")
    sheet.auto_filter.ref = None

    # Save the updated workbook
    workbook.save(pivot_excel_file_path)
    workbook.close()
    print("Pivot excel file saved successfully.")

    # Construct message_string
    message_string = f"{date} {month} {year}, {execution.replace('_', ' ')} Result for {bench.replace('_', ' ')}:"
    message_string_1 = f"{execution.replace('_', ' ')} Result:"
    message_string_2 = f"{auto_flashing.replace('_', ' ')} Result:"
    
    # Call the function to send message to Teams
    send_message_with_zip_to_teams(teams_webhook_url, message_string, message_string_1, message_string_2, overnight_execution_screenshot_path, auto_flashing_screenshot_path)
    
    # Remove Execution_Summary.png
    os.remove(overnight_execution_screenshot_path)
    os.remove(auto_flashing_screenshot_path)

    # Create ZIP folder
    #create_zip_folder(folder_path)

    # Switch back to the previously active window
    #pyautogui.getWindowsWithTitle(active_window_title_before)[0].activate()
    
    input("All tasks completed successfully...")