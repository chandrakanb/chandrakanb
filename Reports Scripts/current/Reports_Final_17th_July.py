import os
import re
import bs4
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import webbrowser
import pyautogui
import pyperclip
import requests
import base64
from io import StringIO
from html2image import Html2Image
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from PIL import Image

def version_details():
    feature = [
        "Creates Folder",
        "Generate Excel Report",
        "Renames Files",
        "Copies Files",
        "Takes Execution Summary Screenshot using Headless Browser\n|\t",
        #--------------------------------------------------------------------------------------------------.
        "For Overnight Executoin:",
        "Adds \"Script Type\" Column to Excel Report",
        "Creates Separate Sheet for \"DryRun2.3\" and \"DryRun2.4\"",
        "Creates Pivot Table for \"DryRun2.3\" and \"DryRun2.4\"",
        "Send message on teams channel with Execution_Summary.png, Summary_Table.png",
        "Copies Message to Clipboard\n|\t",
        #--------------------------------------------------------------------------------------------------.
        "***Need to attach PDF Files",
        "***Need to attach Auto Flashing Summary png",
        "***Clean the code reffering \"Reports_Final_Clean.py\""
    ]
    print("|\tVersion Details:\n|")
    for i, j in enumerate(feature, 1):
        print(f"|\t{i}.\t{j}.")

def read_html_file(execution_report_details):
    html_file = os.path.join(execution_report_details['execution_report_path'], execution_report_details['html_file'])
    df_summary = pd.read_html(html_file)
    df_summary = df_summary[0]
    df_summary = df_summary.filter(df_summary.columns[:2])
    infile = open(html_file)
    lines = infile.readlines()
    FileLines = "".join(lines)
    soup = BeautifulSoup(FileLines, 'html.parser')
    reportframe_add = soup.find('div', attrs={'id':'reportframe'})
    content_src = None
    for content in reportframe_add.contents:
        if isinstance(content, bs4.Tag):
            content_src= content['src']
            break
    content_src = content_src.replace("../","")
    content_src_paths = content_src.split("/")
    curr_path = execution_report_details['execution_report_path']
    curr_path = os.path.normpath(curr_path + os.sep + os.pardir)
    
    curr_path = os.path.join(curr_path,'XMLReport','Segments','XMLReportSegment_0.xml')
    
    tree = ET.parse(curr_path)
    root = tree.getroot()
    mainlist = list()
    for tcs_node in tree.iter("TestCases"):
        for tc_node in tcs_node.iter("TestCase"):
            tc_id = tc_node.find("testCaseID").text
            testCaseObjective = tc_node.find("testCaseObjective").text
            start_time = tc_node.find("startTime").text
            video_logger_link = tc_node.find("VideoLogLink").text
            result = tc_node.find("result").text
            total_time = tc_node.find("totalExecutionTime").text
            end_time = tc_node.find("endTime").text
            
            maindict = {"Test ID": tc_id,"Test Objective":testCaseObjective,"Start Time":start_time,"Video Logger Link":video_logger_link,"Status":result,"Total Time":total_time,"End Time":end_time}
            mainlist.append(maindict)
            
    df_detail_report = pd.DataFrame(mainlist)

    df_detail_report['file_path'] = curr_path
    return df_summary, df_detail_report

def iterate_tuple(dir_tuple):
    for item_tulpe in dir_tuple:
        if isinstance(item_tulpe, tuple):
            iterate_tuple(item_tulpe)
        elif isinstance(item_tulpe, list):
            for html_file in item_tulpe:
                if html_file == "HTMLReport.html":
                    execution_report_path = dir_tuple[0]
                    execution_report_name = execution_report_path.replace(os.getcwd(), "")

                    sub_folder_list = execution_report_name.split(os.sep)
                    test_pack = Iteration_count = None
                    for each_folder in sub_folder_list:
                        if "ITERATION" in each_folder.upper():
                            Iteration_count = each_folder
                            x = re.search("Iteration_[0-9]+", each_folder)
                            Iteration_count = x.group()
                            test_pack = each_folder.replace("_{}".format(Iteration_count),"")
                    Bench_name = sub_folder_list[1]
                    execution_report_name = sub_folder_list[2]
                    return {"execution_report_name": execution_report_name,
                            "execution_report_path": execution_report_path, "Bench_name":Bench_name,
                            "Iteration_count": Iteration_count,
                            "Test_Pack": test_pack,
                            "html_file": html_file}
    return

def add_additional_columns(df, execution_report_details):
    for col_ in ['Bench_name', 'execution_report_name', 'Test_Pack', 'Iteration_count']:
        df[col_] = execution_report_details[col_]
    return df

def generate_Detailed_Report_table(df):
    df_1 = df.filter(["Test_Pack", "Test ID"])
    df['Unique_col'] = df['Test_Pack'] + df['Test ID']
    df['Total Time'] = [datetime.strptime(i, "%H:%M:%S") for i in df['Total Time']]
    df_2 = pd.pivot_table(df, index=["Test_Pack", "Test ID"], columns='Status', aggfunc='count', values='Unique_col')
    return df_2

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            cell[0].value = value

def create_folder(input_string):
    os.makedirs(folder_name, exist_ok=True)
    print(f"|\tFolder created Successfully.")
    print('|' + '_' * 100 + '\n|\t')

def copy_pdf_files(input_string):
    pdf_reports_dir = os.path.join("XMLReports", "PdfReports")
    shutil.copy(os.path.join(pdf_reports_dir, "DashboardReport.pdf"), os.path.join(folder_name, f"{folder_name}_DashboardReport.pdf"))
    shutil.copy(os.path.join(pdf_reports_dir, "DetailedReport.pdf"), os.path.join(folder_name, f"{folder_name}_DetailedReport.pdf"))
    print(f"|\tPDF files copied Successfully.")
    print('|' + '_' * 100 + '\n')

def take_screenshot(html_file_path, screenshot_path, crop_coordinates):
    # Set up Edge options for headless mode
    options = Options()
    options.use_chromium = True
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    
    # Construct the file URL using 'file://' protocol
    file_url = 'file:///' + os.path.abspath(html_file_path).replace('\\', '/')
    
    # Set up the Edge WebDriver with headless options
    driver = webdriver.Edge(options=options)
    
    # Open the HTML file in the headless browser
    driver.get(file_url)
    
    # Take a screenshot of the entire page
    screenshot_data = driver.get_screenshot_as_png()
    
    # Save the screenshot to a temporary file
    temp_screenshot_path = 'temp_screenshot.png'
    with open(temp_screenshot_path, 'wb') as f:
        f.write(screenshot_data)
    
    # Open the screenshot image
    screenshot = Image.open(temp_screenshot_path)
    
    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(crop_coordinates)
    
    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(screenshot_path)
    print('.' + '_' * 100 + '\n|\t')
    print(f"|\tScreenshot Saved Successfully.")
    print('|' + '_' * 100 + '\n|\t')
    
    # Clean up and close the browser
    driver.quit()
    os.remove(temp_screenshot_path)

def crop_table_screenshots(table_image_crop_coordinates, table_image_path):
    # Open the screenshot image
    screenshot = Image.open(table_image_path)
    
    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(table_image_crop_coordinates)
    
    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(table_image_path)

def cut_paste_png_files(folder_name):
    # Assuming the PNG files are in the current directory
    png_files = ['DryRun2.3_Summary.png', 'DryRun2.4_Summary.png']
    
    for png_file in png_files:
        src_path = os.path.join(os.getcwd(), png_file)  # Assuming files are in the current directory
        dest_path = os.path.join(folder_name, png_file)
        
        try:
            shutil.copy(src_path, dest_path)
            os.remove(src_path)
            print(f"|\t\"{png_file}\" moved Successfully.")
            print('|' + '_' * 100 + '\n|\t')
        except FileNotFoundError:
            print(f"|\t\"{png_file}\" not found.")
            print('|' + '_' * 100 + '\n|\t')
        except Exception as e:
            print(f"|\tError while moving \"{png_file}\": {e}")
            print('|' + '_' * 100 + '\n|\t')

def get_date(input_date):
    # Determine suffix based on last digit
    suffix = 'st' if input_date.endswith('1') and not input_date.endswith('11') else \
         'nd' if input_date.endswith('2') and not input_date.endswith('12') else \
         'rd' if input_date.endswith('3') and not input_date.endswith('13') else \
         'th'
         
    # Return ordinal string
    return input_date + suffix

def get_month(input_month):
    # Determine month based input number
    month = 'January' if input_month == '1' else \
         'February' if input_month == '2' else \
         'March' if input_month == '3' else \
         'April' if input_month == '4' else \
         'May' if input_month == '5' else \
         'June' if input_month == '6' else \
         'July' if input_month == '7' else \
         'August' if input_month == '8' else \
         'September' if input_month == '9' else \
         'October' if input_month == '10' else \
         'November' if input_month == '11' else \
         'December' if input_month == '12' else \
         None
         
    # Return month
    return month

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def convert_html_to_image(html_content, output_path, edge_path):
    hti = Html2Image(browser_executable=edge_path)
    hti.screenshot(html_str=html_content, save_as=output_path)

def df_to_html_inline_css(df, font_size, border_width, padding, fixed_width):
    rows = df.to_dict(orient='records')
    headers = df.columns.tolist()
    
    # Define fixed widths for 2nd, 3rd, 4th, and 5th columns
    fixed_widths = {
        1: fixed_width,  # 2nd column
        2: fixed_width,  # 3rd column
        3: fixed_width,  # 4th column
        4: fixed_width   # 5th column
    }
    
    html_table = (
        f"<table border='1' style='border-collapse: collapse; width: auto; font-size: {font_size}px;'>"
    )
    
    # Add headers
    html_table += "<tr>"
    for i, header in enumerate(headers):
        width_style = f"width: {fixed_widths[i]};" if i in fixed_widths else ""
        align_style = "text-align: center;" if i > 0 else ""  # Center align for 2nd to 5th columns
        html_table += f"<td style='border: {border_width}px solid black; padding: {padding}px; {width_style} {align_style} font-weight: bold;'>{header}</td>"
    html_table += "</tr>"
    
    # Add data rows
    for row_idx, row in enumerate(rows):
        html_table += "<tr>"
        for col_idx, cell in enumerate(row.values()):
            width_style = f"width: {fixed_widths[col_idx]};" if col_idx in fixed_widths else ""
            align_style = "text-align: center;" if col_idx > 0 else ""  # Center align for 2nd to 5th columns 
            bold_style = "font-weight: bold;" if row_idx in [3, 3] or col_idx in [0, 4] else ""  # Bold for 1st and 5th rows and columns
            html_table += f"<td style='border: {border_width}px solid black; padding: {padding}px; {width_style} {align_style} {bold_style}'>{cell}</td>"
        html_table += "</tr>"
    
    html_table += "</table><br>"
    
    return html_table

def send_message_to_teams(teams_webhook_url, message_string_1, message_string_2, message_string_3, message_string_4, screenshot_path, table1_image_path, table2_image_path):
    try:
        # Prepare the request headers
        headers = {'Content-Type': 'application/json'}
        
        # Prepare the image attachments
        with open(screenshot_path, "rb") as f:
            encoded_screenshot = base64.b64encode(f.read()).decode('utf-8')
        
        with open(table1_image_path, "rb") as f:
            encoded_table1 = base64.b64encode(f.read()).decode('utf-8')
        
        with open(table2_image_path, "rb") as f:
            encoded_table2 = base64.b64encode(f.read()).decode('utf-8')
        
        # Construct the adaptive card content
        adaptive_card_content = [
            {
                "type": "TextBlock",
                "text": message_string_1,
                "wrap": True
            },
            {
                "type": "Image",
                "url": f"data:image/png;base64,{encoded_screenshot}",
                "size": "Auto"  # Set to "Auto" to maintain original size and high quality
            },
            {
                "type": "TextBlock",
                "text": message_string_2,
                "wrap": True
            },
            {
                "type": "Image",
                "url": f"data:image/png;base64,{encoded_table1}",
                "size": "Auto"
            },
            {
                "type": "TextBlock",
                "text": message_string_3,
                "wrap": True
            },
            {
                "type": "Image",
                "url": f"data:image/png;base64,{encoded_table2}",
                "size": "Auto"
            },
            {
                "type": "TextBlock",
                "text": message_string_4,
                "wrap": True
            }
        ]

        # Prepare the adaptive card payload with the image and tables
        payload = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "content": {
                        "type": "AdaptiveCard",
                        "version": "1.2",
                        "body": adaptive_card_content,
                        "msteams": {
                            "width": "Full"
                        }
                    }
                }
            ]
        }

        # Send the request
        response = requests.post(teams_webhook_url, headers=headers, json=payload)

        # Check for success
        if response.status_code == 200:
            print('.' + '_' * 100 + '\n|\t')
            print(f"|\tMessage sent Successfully to Microsoft Teams!")
            print('|' + '_' * 100 + '\n|\t')
        else:
            print(f"|\tFailed to send message. Status code: {response.status_code}, Response: {response.text}")
            print('|' + '_' * 100 + '\n|\t')

    except Exception as e:
        print(f"|\tAn error occurred: {e}")
        print('|' + '_' * 100 + '\n|\t')

if __name__ == "__main__":
    print('\n|' + '‾' * 100)
    version_details()
    print('|' + '_' * 100 + '\n|\t')
    
    # Get the currently active window title before opening the HTML file
    active_window_title_before = pyautogui.getActiveWindow().title
    
    # Path to the edge executable
    edge_path = r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe'
    
    #Teams webhook URL
    teams_webhook_url = 'https://kpitc.webhook.office.com/webhookb2/0e77225f-5e51-4ff3-bddd-87d2f5da5fc1@3539451e-b46e-4a26-a242-ff61502855c7/IncomingWebhook/bb555a09021747a696f5702d10233e92/fbaa0227-8d16-4af9-aaf0-17c8889162fe'

   # Take Date from User and add correct suffix
    input_date = input("|\tEnter Date: ")
    date = get_date(input_date)
    print('|' + '_' * 100 + '\n|\t')

    # Take Month from User
    input_month = input("|\tEnter month: ")
    month = get_month(input_month)
    print('|' + '_' * 100 + '\n|\t')
    
    # Take Execution Type from User
    execution_type = input("|\tExecution:\n|\t\t1. Overnight Execution\n|\t\t2. Auto Flashing\n|\tEnter your choice (1 or 2): ")
    execution = "Overnight_Execution" if execution_type == "1" else "Auto_Flashing" if execution_type == "2" else None
    print('|' + '_' * 100 + '\n|\t')
    
    # Take Bench Number from User
    input_bench = int(input("|\tEnter Bench Number: "))
    bench = (
        f"Bench0{input_bench}_T_Variant" if input_bench == 3 else
        f"Bench0{input_bench}_R_Variant" if input_bench == 6 else
        f"Bench0{input_bench}_TYAW" if input_bench == 9 else
        f"Bench{input_bench}_3BMA" if input_bench == 10 else
        f"Bench0{input_bench}_Q_Variant"
    )

    print('|' + '_' * 100 + '\n|\t')
    
    # Construct input_string
    parameters = [execution, date, month, bench]
    input_string = '_'.join(parameters) + '_'

    # Folder name
    folder_name = input_string.strip("_").replace(" ", "_")
    
    # Create folder
    create_folder(input_string)
    
    # Copy pdf files
    copy_pdf_files(input_string)
    
    # Take Execution_Summary.png
    html_file_path = 'MainDetailedReport.html'
    screenshot_path = os.path.join(folder_name, f"{folder_name}_Execution_Summary.png")
    crop_coordinates = (90, 63, 1365, 341) # crop coordinates: (left, upper, right, lower)
    
    # Call Screenshot Function 
    take_screenshot(html_file_path, screenshot_path, crop_coordinates)
    
    # Read Excel and HTML files
    df_summary_main = pd.DataFrame([])
    df_detail_report_main = pd.DataFrame([])

    for dir_tuple in os.walk(os.getcwd()):
        if iterate_tuple(dir_tuple):
            execution_report_details = iterate_tuple(dir_tuple)
            try:
                df_summary, df_detail_report = read_html_file(execution_report_details)
                df_summary = df_summary.T
                df_summary.columns = df_summary.loc['Summary'].to_list()
                df_summary = add_additional_columns(df_summary, execution_report_details)
                df_detail_report = add_additional_columns(df_detail_report, execution_report_details)
                df_summary_main = pd.concat([df_summary_main, df_summary], axis=0) if not df_summary_main.empty else df_summary
                df_detail_report_main = pd.concat([df_detail_report_main, df_detail_report], axis=0) if not df_detail_report_main.empty else df_detail_report
            except ValueError:
                print(execution_report_details)

    df_summary_main = df_summary_main.loc[df_summary_main['Start Time'] != 'Start Time']
    df_detail_report_main['Start Time_1'] = [datetime.strptime(i, "%d %b %Y %H:%M:%S") for i in df_detail_report_main['Start Time']]
    df_detail_report_main['Time_1'] = [datetime.strptime(i, "%H:%M:%S") for i in df_detail_report_main['Total Time']]
    End_Time = []
    for x, y in zip(df_detail_report_main['Start Time_1'], df_detail_report_main['Time_1']):
        __day__ = 0 if y.year == 1900 else y.day
        End_Time.append(x + timedelta(days=__day__, hours=y.hour, minutes=y.minute, seconds=y.second))
    df_detail_report_main['End Time'] = End_Time

    df_detail_report_main = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name",
                                                        "Iteration_count", "file_path"])

    df_detail_report_main_table = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name"])

    df_detail_report_main_table['Remark Failure Reason'] = ""

    # Write to Excel
    excel_file_path = os.path.join(folder_name, f"{folder_name}_Execution_Report.xlsx")
    writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')
    df_detail_report_main_table.to_excel(writer, sheet_name='Detailed Report', index=False)
    writer._save()
    print(f"|\tExcel file Saved Successfully.")
    print('|' + '_' * 100 + '\n|\t')
    
    #Define the border style and bold text
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    if execution == "Overnight_Execution":
        # Process Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        sheet.insert_cols(2)
        cell = sheet.cell(row=1, column=2)
        cell.value = "Script Type"
        cell.font = Font(bold=True)
        cell.border = thin_border
        apply_filter(sheet, "KPIT", "C")
        update_column_B(sheet, "KPIT", "Test Script")
        sheet.auto_filter.ref = None
        apply_filter(sheet, "pre", "C")
        update_column_B(sheet, "pre", "Precondition")
        sheet.auto_filter.ref = None
        apply_filter(sheet, "post", "C")
        update_column_B(sheet, "post", "Postcondition")
        sheet.auto_filter.ref = None
        apply_filter(sheet, "auto", "C")
        update_column_B(sheet, "auto", "Precondition")
        sheet.auto_filter.ref = None
        workbook.save(excel_file_path)
    
        # Load the Excel workbook and the sheet with the result data
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # Apply filter and create new sheets, including the first column (headers) in bold and bordered
        for version in ["2.3", "2.4"]:
            ws.auto_filter.ref = ws.dimensions
            apply_filter(ws, version, 'A')
            wb.create_sheet(f"DryRun{version}")
            new_sheet = wb[f"DryRun{version}"]
        
            # Copy headers
            for cell in ws[1]:
                new_sheet[cell.coordinate].value = cell.value
                new_sheet[cell.coordinate].font = bold_font
                new_sheet[cell.coordinate].border = thin_border

            # Copy filtered rows    
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if version in str(row[0]):
                    new_sheet.append(row)
            ws.auto_filter.ref = None
        
            # Create pivot table structure in new summary sheet
            summary_sheet_name = f"DryRun{version} Summary"
            wb.create_sheet(summary_sheet_name)
            summary_sheet = wb[summary_sheet_name]

            # Add headers to summary sheet
            headers = ["Script Type", "Fail", "Pass", "Verify Manually", "Total"]
            for col_num, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = bold_font
                #cell.border = thin_border

            # Prepare data for pivot table
            data = new_sheet.values
            df = pd.DataFrame(data)
            df.columns = df.iloc[0]
            df = df[1:]
            pivot = pd.pivot_table(df, index=["Script Type"], columns=["Status"], aggfunc='size', fill_value=0).reset_index()
        
            # Convert columns to numeric to avoid type errors
            for col in ["Fail", "Pass", "Verify Manually"]:
                if col in pivot.columns:
                    pivot[col] = pd.to_numeric(pivot[col], errors='coerce').fillna(0).astype(int)
                else:
                    pivot[col] = 0

            pivot['Total'] = pivot[["Fail", "Pass", "Verify Manually"]].sum(axis=1)
        
            # Reorder columns
            pivot = pivot[["Script Type", "Fail", "Pass", "Verify Manually", "Total"]]
        
            # Add total row
            total_row = pivot.sum(numeric_only=True)
            total_row["Script Type"] = "Total"
            pivot = pd.concat([pivot, total_row.to_frame().T], ignore_index=True)
        
            # Write pivot table to summary sheet
            for r_idx, row in enumerate(dataframe_to_rows(pivot, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = summary_sheet.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = bold_font if r_idx == 1 else None
                    cell.border = thin_border
        
            # Bold "Script Type" column
            for cell in summary_sheet["A:A"]:
                cell.font = bold_font
        
            # Access the DryRun2.x Summary sheet
            pivot_sheet_name = f"DryRun{version} Summary"
            pivot_sheet = wb[pivot_sheet_name]

            # Prepare data for pivot table
            data = pivot_sheet.values
            columns = next(data)  # Fetch the first row for column headers

            # Convert to DataFrame
            df = pd.DataFrame(data, columns=columns)
            font_size = int(20)
            border_width = int(2)
            padding  = int(2)
            fixed_width = "80px"
            # Convert DataFrame to HTML with custom styling
            html_table = df_to_html_inline_css(df, font_size, border_width, padding, fixed_width)
        
            # Generate Pandas DataFrame from HTML
            if version == "2.3":
                HTML_2_3 = html_table
            elif version == "2.4":
                HTML_2_4 = html_table
        
        # Save the updated workbook
        wb.save(excel_file_path)
        wb.close()
    
        print(f"|\tExcel file Updated Successfully.")
        print('|' + '_' * 100 + '\n')
        
        # Convert HTML tables to images
        table1_image_path = 'DryRun2.3_Summary.png'
        table2_image_path = 'DryRun2.4_Summary.png'
        table_image_crop_coordinates = (5, 5, 485, 185) # crop coordinates: (left, upper, right, lower)
        
        convert_html_to_image(HTML_2_3, table1_image_path, edge_path)
        crop_table_screenshots(table_image_crop_coordinates, table1_image_path)
        convert_html_to_image(HTML_2_4, table2_image_path, edge_path)
        crop_table_screenshots(table_image_crop_coordinates, table2_image_path)

        # Construct message_string
        # date_month = f"{date} {month}, "
        overnight_execution_result = "Overnight Execution Result:"
        flashing_result = "Auto Flashing Result:"
        DryRun2_3 = "DryRun2.3"
        DryRun2_4 = "DryRun2.4"
        if True:
            message_string_1 = (
                f"{date} {month}, {execution.replace('_', ' ')} {bench.replace('_', ' ')}:\n\n"
                f"{overnight_execution_result}\n\n"
            )
            message_string_2 = (
                f"{DryRun2_3} {overnight_execution_result}\n\n"
            )
            message_string_3 = (
                f"{DryRun2_4} {overnight_execution_result}\n\n"
            )
            message_string_4 = (
                f"{flashing_result}"
            )

        # Call the function to send message to Teams
        send_message_to_teams(teams_webhook_url, message_string_1, message_string_2, message_string_3, message_string_4, screenshot_path, table1_image_path, table2_image_path)

        # Copy the output string to the clipboard
        message_string = (
            f"{date} {month}, {execution.replace('_', ' ')} {bench.replace('_', ' ')}:\n\n"
            f"{overnight_execution_result}\n\n"
            f"{DryRun2_3} {overnight_execution_result}\n\n"
            f"{DryRun2_4} {overnight_execution_result}\n\n"
            f"{flashing_result}"
        )
        pyperclip.copy(message_string)
        output_message = pyperclip.paste()
        if output_message == message_string:
            print("|\tMessage copied Successfully.")
            print('|' + '_' * 100 + '\n|\t')
        
        # Cut paste png files
        cut_paste_png_files(folder_name)

    # Switch back to the previously active window
    pyautogui.getWindowsWithTitle(active_window_title_before)[0].activate()
    
    print("|\tAll tasks completed Successfully.")
    print('|' + '_' * 100 + '\n|\t')
    input("|\tPress Enter to exit...")
    print('|' + '_' * 100 + '\n')