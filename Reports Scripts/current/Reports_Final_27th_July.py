import os
import re
import bs4
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import webbrowser
import pyautogui
import pyperclip
import requests
import base64
from io import StringIO
from html2image import Html2Image
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from PIL import Image
import subprocess
from contextlib import contextmanager, redirect_stdout, redirect_stderr

def read_html_file(execution_report_details):
    html_file = os.path.join(execution_report_details['execution_report_path'], execution_report_details['html_file'])
    df_summary = pd.read_html(html_file)
    df_summary = df_summary[0]
    df_summary = df_summary.filter(df_summary.columns[:2])
    infile = open(html_file)
    lines = infile.readlines()
    FileLines = "".join(lines)
    soup = BeautifulSoup(FileLines, 'html.parser')
    reportframe_add = soup.find('div', attrs={'id':'reportframe'})
    content_src = None
    for content in reportframe_add.contents:
        if isinstance(content, bs4.Tag):
            content_src= content['src']
            break
    content_src = content_src.replace("../","")
    content_src_paths = content_src.split("/")
    curr_path = execution_report_details['execution_report_path']
    curr_path = os.path.normpath(curr_path + os.sep + os.pardir)
    
    curr_path = os.path.join(curr_path,'XMLReport','Segments','XMLReportSegment_0.xml')
    
    tree = ET.parse(curr_path)
    root = tree.getroot()
    mainlist = list()
    for tcs_node in tree.iter("TestCases"):
        for tc_node in tcs_node.iter("TestCase"):
            tc_id = tc_node.find("testCaseID").text
            testCaseObjective = tc_node.find("testCaseObjective").text
            start_time = tc_node.find("startTime").text
            video_logger_link = tc_node.find("VideoLogLink").text
            result = tc_node.find("result").text
            total_time = tc_node.find("totalExecutionTime").text
            end_time = tc_node.find("endTime").text
            
            maindict = {"Test ID": tc_id,"Test Objective":testCaseObjective,"Start Time":start_time,"Video Logger Link":video_logger_link,"Status":result,"Total Time":total_time,"End Time":end_time}
            mainlist.append(maindict)
            
    df_detail_report = pd.DataFrame(mainlist)

    df_detail_report['file_path'] = curr_path
    return df_summary, df_detail_report

def iterate_tuple(dir_tuple):
    for item_tulpe in dir_tuple:
        if isinstance(item_tulpe, tuple):
            iterate_tuple(item_tulpe)
        elif isinstance(item_tulpe, list):
            for html_file in item_tulpe:
                if html_file == "HTMLReport.html":
                    execution_report_path = dir_tuple[0]
                    execution_report_name = execution_report_path.replace(os.getcwd(), "")

                    sub_folder_list = execution_report_name.split(os.sep)
                    test_pack = Iteration_count = None
                    for each_folder in sub_folder_list:
                        if "ITERATION" in each_folder.upper():
                            Iteration_count = each_folder
                            x = re.search("Iteration_[0-9]+", each_folder)
                            Iteration_count = x.group()
                            test_pack = each_folder.replace("_{}".format(Iteration_count),"")
                    Bench_name = sub_folder_list[1]
                    execution_report_name = sub_folder_list[2]
                    return {"execution_report_name": execution_report_name,
                            "execution_report_path": execution_report_path, "Bench_name":Bench_name,
                            "Iteration_count": Iteration_count,
                            "Test_Pack": test_pack,
                            "html_file": html_file}
    return

def add_additional_columns(df, execution_report_details):
    for col_ in ['Bench_name', 'execution_report_name', 'Test_Pack', 'Iteration_count']:
        df[col_] = execution_report_details[col_]
    return df

def generate_Detailed_Report_table(df):
    df_1 = df.filter(["Test_Pack", "Test ID"])
    df['Unique_col'] = df['Test_Pack'] + df['Test ID']
    df['Total Time'] = [datetime.strptime(i, "%H:%M:%S") for i in df['Total Time']]
    df_2 = pd.pivot_table(df, index=["Test_Pack", "Test ID"], columns='Status', aggfunc='count', values='Unique_col')
    return df_2

def apply_filter(sheet, keyword, column):
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            cell[0].value = value

def create_folder(folder_name):
    """Create a directory if it does not exist."""
    os.makedirs(folder_name, exist_ok=True)
    print(f"|\tFolder created Successfully.")
    print('|' + '_' * 100 + '\n|\t')

def copy_pdf_files(folder_name):
    pdf_reports_dir = os.path.join("XMLReports", "PdfReports")
    shutil.copy(os.path.join(pdf_reports_dir, "DashboardReport.pdf"), os.path.join(folder_name, f"{folder_name}_DashboardReport.pdf"))
    shutil.copy(os.path.join(pdf_reports_dir, "DetailedReport.pdf"), os.path.join(folder_name, f"{folder_name}_DetailedReport.pdf"))
    print(f"|\tPDF files copied Successfully.")
    print('|' + '_' * 100 + '\n')

def take_screenshot(html_file_path, screenshot_path, crop_coordinates):
    """Take a screenshot of the specified HTML file and save it."""
    # Set up Edge options for headless mode
    options = Options()
    options.use_chromium = True
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    
    # Construct the file URL using 'file://' protocol
    file_url = 'file:///' + os.path.abspath(html_file_path).replace('\\', '/')
    
    # Set up the Edge WebDriver with headless options
    driver = webdriver.Edge(options=options)
    
    # Open the HTML file in the headless browser
    driver.get(file_url)
    
    # Take a screenshot of the entire page
    screenshot_data = driver.get_screenshot_as_png()
    
    # Save the screenshot to a temporary file
    temp_screenshot_path = 'temp_screenshot.png'
    with open(temp_screenshot_path, 'wb') as f:
        f.write(screenshot_data)
    
    # Open the screenshot image
    screenshot = Image.open(temp_screenshot_path)
    
    # Crop the screenshot based on the provided coordinates
    cropped_screenshot = screenshot.crop(crop_coordinates)
    
    # Save the cropped screenshot to the specified path
    cropped_screenshot.save(screenshot_path)
    print('.' + '_' * 100 + '\n|\t')
    print(f"|\tScreenshot Saved Successfully.")
    print('|' + '_' * 100 + '\n|\t')
    
    # Clean up and close the browser
    driver.quit()
    os.remove(temp_screenshot_path)

def get_execution_date_month():
    """
    This function reads an HTML report file to extract the start time of the execution.
    It then parses the start time to obtain the date, month, year, hours, and minutes.
    The date is adjusted if the hour is between 12 midnight and 12 noon of the next day.
    Finally, it converts the abbreviated month name to the full month name and returns the date and month.
    """
    # Read the content of the HTML file
    with open('MainDetailedReport.html', 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the start time
    start_time_tag = soup.find('th', string='Start Time')
    start_time_value = start_time_tag.find_next_sibling('td').text if start_time_tag else None

    # Split the start time into date, month, year, hours, and minutes
    date_parts = start_time_value.split(' ')
    date_digit = int(date_parts[0])  # "date" as an integer
    month_abbr = date_parts[1]  # "month"
    year = date_parts[2]  # "year"
    time_parts = date_parts[3].split(':')
    time_hh = time_parts[0]  # "hours"
    time_mm = time_parts[1]  # "minutes"

    # Adjust the date based on the hour
    # If hours is between 12 midnight and next day's 12 noon, then date minus one
    if int(time_hh) <= 12:
        date_digit -= 1
    
    # Add correct suffix to the date based on last digit
    date = get_date(date_digit)
    
    # Convert abbreviated month to full month name
    month = get_month(month_abbr)

    return date, month, year

def get_date(date_digit):
    """Return the date with the appropriate ordinal suffix."""
    date_str = str(date_digit)
    suffix = 'st' if date_str.endswith('1') and not date_str.endswith('11') else \
             'nd' if date_str.endswith('2') and not date_str.endswith('12') else \
             'rd' if date_str.endswith('3') and not date_str.endswith('13') else \
             'th'
    return date_str + suffix

def get_month(month_abbr):
    """Return the full month name based on abbreviated month names."""
    months = {
        "Jan": "January", "Feb": "February", "Mar": "March",
        "Apr": "April", "May": "May", "Jun": "June",
        "Jul": "July", "Aug": "August", "Sep": "September",
        "Oct": "October", "Nov": "November", "Dec": "December"
    }
    return months.get(month_abbr, month_abbr)

def get_bench(input_bench):
    """Return the bench name with the appropriate variant name."""
    bench = (
        f"Bench0{input_bench}_T_Variant" if input_bench == 7 else
        f"Bench0{input_bench}_R_Variant" if input_bench == 6 else
        f"Bench0{input_bench}_TYAW" if input_bench == 9 else
        f"Bench{input_bench}_3BMA" if input_bench == 10 else
        f"Bench0{input_bench}_Q_Variant"
    )
    return bench

def send_message_to_teams(teams_webhook_url, message_string, screenshot_path):
    try:
        # Prepare the request headers
        headers = {'Content-Type': 'application/json'}
        
        # Prepare the image attachments
        with open(screenshot_path, "rb") as f:
            encoded_screenshot = base64.b64encode(f.read()).decode('utf-8')
        
        # Construct the adaptive card content
        adaptive_card_content = [
            {
                "type": "TextBlock",
                "text": message_string,
                "wrap": True
            },
            {
                "type": "Image",
                "url": f"data:image/png;base64,{encoded_screenshot}",
                "size": "Auto"  # Set to "Auto" to maintain original size and high quality
            }
        ]

        # Prepare the adaptive card payload with the image and tables
        payload = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "content": {
                        "type": "AdaptiveCard",
                        "version": "1.2",
                        "body": adaptive_card_content,
                        "msteams": {
                            "width": "Full"
                        }
                    }
                }
            ]
        }

        # Send the request
        response = requests.post(teams_webhook_url, headers=headers, json=payload)

        # Check for success
        if response.status_code == 200:
            print(f"|\tMessage sent Successfully to Microsoft Teams!")
            print('|' + '_' * 100 + '\n|\t')
        else:
            print(f"|\tFailed to send message. Status code: {response.status_code}, Response: {response.text}")
            print('|' + '_' * 100 + '\n|\t')

    except Exception as e:
        print(f"|\tAn error occurred: {e}")
        print('|' + '_' * 100 + '\n|\t')

if __name__ == "__main__":
    # Get the currently active window title before opening the HTML file
    active_window_title_before = pyautogui.getActiveWindow().title
    
    #Teams webhook URL
    teams_webhook_url = 'https://kpitc.webhook.office.com/webhookb2/0e77225f-5e51-4ff3-bddd-87d2f5da5fc1@3539451e-b46e-4a26-a242-ff61502855c7/IncomingWebhook/bb555a09021747a696f5702d10233e92/fbaa0227-8d16-4af9-aaf0-17c8889162fe'

    # Get the execution date and month from the HTML report
    date, month, year = get_execution_date_month()
    
    # Prompt the user to enter the bench number
    input_bench = int(input("Enter Bench Number: "))
    bench = get_bench(input_bench)
    print('|' + '_' * 100 + '\n|\t')
    
    # Set the execution type
    execution = "Overnight_Execution"
    print('|' + '_' * 100 + '\n|\t')
    
    # Construct folder_name
    parameters = [execution, date, month, bench]
    folder_name = '_'.join(parameters).strip("_").replace(" ", "_")

    # Create folder
    create_folder(folder_name)
    
    # Copy pdf files
    copy_pdf_files(folder_name)
    
    # Take Execution_Summary.png
    html_file_path = 'MainDetailedReport.html'
    screenshot_path = os.path.join(folder_name, f"{folder_name}_Summary_Table_Pie_Chart.png")
    crop_coordinates = (90, 63, 1365, 341) # crop coordinates: (left, upper, right, lower)
    
    # Call Screenshot Function 
    take_screenshot(html_file_path, screenshot_path, crop_coordinates)
    
    # Read Excel and HTML files
    df_summary_main = pd.DataFrame([])
    df_detail_report_main = pd.DataFrame([])

    for dir_tuple in os.walk(os.getcwd()):
        if iterate_tuple(dir_tuple):
            execution_report_details = iterate_tuple(dir_tuple)
            try:
                df_summary, df_detail_report = read_html_file(execution_report_details)
                df_summary = df_summary.T
                df_summary.columns = df_summary.loc['Summary'].to_list()
                df_summary = add_additional_columns(df_summary, execution_report_details)
                df_detail_report = add_additional_columns(df_detail_report, execution_report_details)
                df_summary_main = pd.concat([df_summary_main, df_summary], axis=0) if not df_summary_main.empty else df_summary
                df_detail_report_main = pd.concat([df_detail_report_main, df_detail_report], axis=0) if not df_detail_report_main.empty else df_detail_report
            except ValueError:
                print(execution_report_details)

    df_summary_main = df_summary_main.loc[df_summary_main['Start Time'] != 'Start Time']
    df_detail_report_main['Start Time_1'] = [datetime.strptime(i, "%d %b %Y %H:%M:%S") for i in df_detail_report_main['Start Time']]
    df_detail_report_main['Time_1'] = [datetime.strptime(i, "%H:%M:%S") for i in df_detail_report_main['Total Time']]
    End_Time = []
    for x, y in zip(df_detail_report_main['Start Time_1'], df_detail_report_main['Time_1']):
        __day__ = 0 if y.year == 1900 else y.day
        End_Time.append(x + timedelta(days=__day__, hours=y.hour, minutes=y.minute, seconds=y.second))
    df_detail_report_main['End Time'] = End_Time

    df_detail_report_main = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name",
                                                        "Iteration_count", "file_path"])

    df_detail_report_main_table = df_detail_report_main.filter(["Test_Pack","Test ID", "Status", "Total Time", "Bench_name"])

    df_detail_report_main_table['Remark Failure Reason'] = ""

    # Write to Excel
    excel_file_path = os.path.join(folder_name, f"{folder_name}_Execution_Report.xlsx")
    writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')
    df_detail_report_main_table.to_excel(writer, sheet_name='Detailed Report', index=False)
    writer._save()
    print(f"|\tExcel file Saved Successfully.")
    print('|' + '_' * 100 + '\n|\t')
    
    #Define the border style and bold text
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    # Process Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    sheet.insert_cols(2)
    cell = sheet.cell(row=1, column=2)
    cell.value = "Script Type"
    cell.font = Font(bold=True)
    cell.border = thin_border
    apply_filter(sheet, "KPIT", "C")
    update_column_B(sheet, "KPIT", "Test Script")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "pre", "C")
    update_column_B(sheet, "pre", "Precondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "post", "C")
    update_column_B(sheet, "post", "Postcondition")
    sheet.auto_filter.ref = None
    apply_filter(sheet, "auto", "C")
    update_column_B(sheet, "auto", "Precondition")
    sheet.auto_filter.ref = None

    # Save the updated workbook
    workbook.save(excel_file_path)
    workbook.close()    
    print(f"|\tExcel file Updated Successfully.")
    print('|' + '_' * 100 + '\n|\t')

    # Construct message_string
    message_string = f"{date} {month} {year}, {execution.replace('_', ' ')} Result for {bench.replace('_', ' ')}:\n"

    # Call the function to send message to Teams
    send_message_to_teams(teams_webhook_url, message_string, screenshot_path)

    # Switch back to the previously active window
    pyautogui.getWindowsWithTitle(active_window_title_before)[0].activate()
    
    print("|\tAll tasks completed Successfully.")
    print('|' + '_' * 100 + '\n|\t')
    input("|\tPress Enter to exit...")
    print('|' + '_' * 100 + '\n')