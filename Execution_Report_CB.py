import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def apply_filter(sheet, keyword, column):
    # Apply filter for the keyword in the specified column
    for cell in sheet[column]:
        if isinstance(keyword, str):
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [f"*{keyword}*"])
        else:
            sheet.auto_filter.add_filter_column(cell.col_idx - 1, [keyword])

def update_column_B(sheet, keyword, value):
    # Iterate through each cell in column B (from B2 to the last row)
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        # Check if the corresponding cell in column C has a value
        if cell[0].offset(column=1).value and isinstance(keyword, str) and keyword.upper() in str(cell[0].offset(column=1).value).upper():
            # Set the value of the current cell in column B
            cell[0].value = value

def create_pivot_table(excel_file_path):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file_path)
        
        # Get the active sheet
        sheet = wb.active
        
        # Create a new sheet for the pivot table
        pivot_sheet = wb.create_sheet(title="temp")
        
        # Create a pivot table
        pivot, pivot_rows, pivot_columns = pivot_table(sheet)
        
        # Add pivot table to the new sheet
        for r_idx, row in enumerate(pivot_rows, 1):
            for c_idx, value in enumerate(row, 1):
                pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Create a new sheet for the transposed pivot table
        transposed_pivot_sheet = wb.create_sheet(title="Pivot_Table")
        
        # Transpose the pivot table
        transposed_pivot = pivot.transpose()
        
        # Add transposed pivot table to the new sheet
        for r_idx, row in enumerate(dataframe_to_rows(transposed_pivot, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                transposed_pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Remove unnecessary rows
        transposed_pivot_sheet.delete_rows(1, 2)
        
        # Replace text at A1 with "Script Type"
        transposed_pivot_sheet['A1'] = "Script Type"
        
        # Remove the "temp" sheet
        wb.remove(wb["temp"])
        
        # Save the workbook with the specified name format
        output_file_name = f"Execution_Report_{pd.Timestamp.now().strftime('%dth_%B')}.xlsx"
        wb.save(output_file_name)
        
        print(f"Created the pivot table and transposed pivot table, saved as '{output_file_name}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def pivot_table(sheet):
    data = sheet.values
    cols = next(data)
    data = list(data)
    idx = [col for col in cols if col == "Script Type"]
    status_index = [col for col in cols if col == "Status"]
    script_type_index = cols.index("Script Type")

    # Get indices for columns
    idx_idx = cols.index(idx[0])
    status_idx = cols.index(status_index[0])
    script_type_idx = cols.index("Script Type")
    
    # Initialize dictionary to store counts
    counts = {}
    for row in data:
        if row[script_type_idx] not in counts:
            counts[row[script_type_idx]] = {}
        if row[status_idx] not in counts[row[script_type_idx]]:
            counts[row[script_type_idx]][row[status_idx]] = 0
        counts[row[script_type_idx]][row[status_idx]] += 1
    
    # Create dataframe from counts
    df = pd.DataFrame(counts).fillna(0)
    
    # Add Total row and column
    df.loc['Total'] = df.sum()
    df['Total'] = df.sum(axis=1)
    
    # Reset index to make it a regular column
    df.reset_index(inplace=True)
    
    return df, dataframe_to_rows(df, index=False, header=True), cols

if __name__ == "__main__":
    excel_file_path = "Execution_Report.xlsx"
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the active sheet
    sheet = workbook.active

    # Insert a column after the 1st column and name it "Script Type"
    sheet.insert_cols(2)
    sheet.cell(row=1, column=2).value = "Script Type"

    # Apply filter for "KPIT" in the 3rd row
    apply_filter(sheet, "KPIT", "C")

    # Update column B for "KPIT"
    update_column_B(sheet, "KPIT", "Test Script")

    # Remove the filter
    sheet.auto_filter.ref = None

    # Apply filter for "pre" in the 3rd row
    apply_filter(sheet, "pre", "C")

    # Update column B for "pre"
    update_column_B(sheet, "pre", "Precondition")

    # Remove the filter
    sheet.auto_filter.ref = None

    # Apply filter for "post" in the 3rd row
    apply_filter(sheet, "post", "C")

    # Update column B for "post"
    update_column_B(sheet, "post", "Postcondition")

    # Remove the filter
    sheet.auto_filter.ref = None

    # Save the changes to a new Excel file named 'Execution_Report_Temp.xlsx'
    workbook.save('Execution_Report_Temp.xlsx')

    # Create pivot table
    create_pivot_table('Execution_Report_Temp.xlsx')

    # Delete the 'Execution_Report_Temp.xlsx' file before exiting the script
    try:
        os.remove('Execution_Report_Temp.xlsx')
        print("                                     ")
    except FileNotFoundError:
        print("                                     ")
